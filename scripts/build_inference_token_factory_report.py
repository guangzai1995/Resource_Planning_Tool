"""Generate the inference token factory leadership report workbook."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from statistics import mean
from typing import Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


REPORT_VERSION = "v2.0"
OUTPUT_FILENAME = "推理token工厂汇报大纲_汇报版_v2.0.xlsx"

STATUS_LABELS = [
    "已验证",
    "有结论但缺系统数据",
    "待验证",
    "已具备无需独立测试",
]

SHEET_SPECS = [
    {"title": "00_目录与版本", "owner": "孙力光", "status": "已验证"},
    {"title": "01_背景与目标", "owner": "孙力光", "status": "已验证"},
    {"title": "02_总览结论", "owner": "孙力光", "status": "已验证"},
    {"title": "03_技术路线地图", "owner": "孙力光", "status": "已验证"},
    {"title": "04_模型量化", "owner": "孙力光", "status": "已验证"},
    {"title": "05_KVCache量化", "owner": "胡正升", "status": "有结论但缺系统数据"},
    {"title": "06_Prefix_KV命中", "owner": "胡正升", "status": "有结论但缺系统数据"},
    {"title": "07_投机解码", "owner": "胡正升", "status": "待验证"},
    {"title": "08_PD分离", "owner": "孙力光", "status": "待验证"},
    {"title": "09_MOE专家并行", "owner": "孙力光", "status": "待验证"},
    {"title": "10_连续批处理", "owner": "孙力光", "status": "已具备无需独立测试"},
    {"title": "11_汇总建议与资源计划", "owner": "孙力光", "status": "待验证"},
    {"title": "12_数据附录", "owner": "孙力光", "status": "已验证"},
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)

TECH_ROWS = [
    ("模型量化", "已验证", "H200 FP8/BF16 已有数据，重点展示吞吐和 TPOT 收益"),
    ("KV Cache 量化", "有结论但缺系统数据", "fp8 可节省 KV 显存，fp4 因 H200 不支持暂不测"),
    ("Prefix/KV 命中", "有结论但缺系统数据", "需补测 prefix_ratio、命中率和 TTFT 下降比例"),
    ("投机解码", "待验证", "候选 token 数 1/3/5，关注 TPOT、接受率和高并发衰减"),
    ("PD 分离", "待验证", "覆盖 H200 P+D、H200 P/H200 D、H200 P/H20 D"),
    ("MOE 专家并行", "待验证", "按 GLM5.1/GLM5.2 FP8 部署口径申请资源"),
    ("连续批处理", "已具备无需独立测试", "框架原生支持，作为能力现状说明"),
]

PRINCIPLE_DIAGRAMS = {
    "04_模型量化": ["BF16/FP16 权重与激活", "FP8/量化模型", "显存下降", "吞吐提升"],
    "05_KVCache量化": ["KV Cache Block", "fp16/bf16 KV", "fp8 KV", "fp4 H200 不支持"],
    "06_Prefix_KV命中": ["共享前缀", "Prefix Hash", "复用 KV", "TTFT 降低"],
    "07_投机解码": ["Draft 候选 token", "Verify", "Accept/Reject", "TPOT 降低"],
    "08_PD分离": ["Prefill(H200)", "KV 传输", "Decode(H200/H20)", "H200 P + H20 D"],
    "09_MOE专家并行": ["Router", "多 GPU Experts", "All2All 聚合", "GLM5.1/GLM5.2 FP8"],
    "10_连续批处理": ["Scheduler", "新请求 Prefill", "旧请求 Decode", "框架原生支持"],
}


def _style_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _style_table_header(ws, row: int, max_col: int, min_col: int = 1) -> None:
    for row_cells in ws.iter_rows(min_row=row, max_row=row, min_col=min_col, max_col=max_col):
        for cell in row_cells:
            cell.fill = SUB_FILL
            cell.font = Font(bold=True)


def _write_chart_source_table(
    ws,
    start_row: int,
    start_col: int,
    headers: list[str],
    rows: list[list[str | float | int]],
) -> int:
    for col_idx, header in enumerate(headers, start=start_col):
        ws.cell(row=start_row, column=col_idx, value=header)

    for row_idx, row_values in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row_values, start=start_col):
            ws.cell(row=row_idx, column=col_idx, value=value)

    last_row = start_row + len(rows)
    max_col = start_col + len(headers) - 1
    _style_range(ws, start_row, last_row, start_col, max_col)
    _style_table_header(ws, start_row, max_col, min_col=start_col)
    return last_row


def _write_header(ws, spec: dict[str, str]) -> None:
    ws["A1"] = "推理 Token 工厂汇报"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws.merge_cells("A1:H1")
    fields = [
        ("报告版本", REPORT_VERSION),
        ("责任人", spec["owner"]),
        ("当前状态", spec["status"]),
        ("数据来源", "见 12_数据附录"),
    ]
    for idx, (label, value) in enumerate(fields, start=2):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=value)
    ws["A4"] = "当前状态"
    ws.freeze_panes = "A6"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 18


def _write_sheet_index(ws) -> None:
    ws.append([])
    ws.append(["Sheet", "名称", "责任人", "状态", "说明"])
    header_row = ws.max_row
    for index, spec in enumerate(SHEET_SPECS):
        ws.append([index, spec["title"], spec["owner"], spec["status"], "按规格生成"])
    _style_range(ws, header_row, ws.max_row, 1, 5)
    _style_table_header(ws, header_row, 5)


def _write_summary(ws) -> None:
    ws.append([])
    ws.append(["技术点", "状态", "汇报口径"])
    header_row = ws.max_row
    for row in TECH_ROWS:
        ws.append(list(row))
    _style_range(ws, header_row, ws.max_row, 1, 3)
    _style_table_header(ws, header_row, 3)


def _write_principle_diagram(ws, labels: list[str]) -> None:
    ws["A6"] = "V2.2 方案：原理简图"
    ws["A6"].font = Font(bold=True, color="1F4E78")
    start_row = 8
    start_col = 1
    for idx, label in enumerate(labels):
        col = start_col + idx * 2
        ws.cell(start_row, col, label)
        ws.cell(start_row, col).fill = SUB_FILL
        ws.cell(start_row, col).font = Font(bold=True)
        ws.cell(start_row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(start_row, col).border = THIN_BORDER
        ws.column_dimensions[ws.cell(start_row, col).column_letter].width = 20
        if idx < len(labels) - 1:
            ws.cell(start_row, col + 1, "→")
            ws.cell(start_row, col + 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(start_row, col + 1).border = THIN_BORDER
            ws.column_dimensions[ws.cell(start_row, col + 1).column_letter].width = 6
    ws["A10"] = "说明"
    ws["A10"].font = Font(bold=True)
    ws["A10"].alignment = Alignment(vertical="center", wrap_text=True)
    ws["A10"].border = THIN_BORDER
    ws["B10"] = "该图用于快速解释技术机制，详细参数与测试设计见下方表格。"
    ws["B10"].alignment = Alignment(vertical="center", wrap_text=True)
    ws["B10"].border = THIN_BORDER
    ws.merge_cells("B10:G10")


def _write_section_title(ws, row: int, title: str) -> int:
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(bold=True, color="1F4E78")
    ws.cell(row=row, column=1).fill = SUB_FILL
    ws.cell(row=row, column=1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    return row + 1


def _append_table(ws, start_row: int, headers: list[str], rows: list[list[str]]) -> int:
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=header)

    for row_idx, row_values in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    last_row = start_row + len(rows)
    max_col = len(headers)
    _style_range(ws, start_row, last_row, 1, max_col)
    _style_table_header(ws, start_row, max_col)
    for row in ws.iter_rows(min_row=start_row, max_row=last_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    return last_row + 1


def _format_wan(value: float | int) -> str:
    return f"{float(value) / 10_000:.1f} 万"


def _format_yi_from_billion(value: float | int) -> str:
    return f"{float(value) * 10:.1f} 亿"


def _format_pct(value: float | int) -> str:
    return f"{float(value):.1f}%"


def _write_background(ws, context: dict[str, object]) -> None:
    start_row = ws.max_row + 2
    row = _write_section_title(ws, start_row, "汇报背景与目标")
    request_scale = _format_wan(context["total_requests"])
    token_scale = _format_yi_from_billion(context["total_tokens_billion"])
    input_ratio = _format_pct(float(context["input_token_ratio"]) * 100.0)
    long_context_ratio = _format_pct(float(context["long_context_ratio"]) * 100.0)
    row = _append_table(
        ws,
        row,
        ["主题", "已有结论/历史数据", "领导汇报口径"],
        [
            [
                "业务定位",
                "工服场景 token 工厂，以推理请求、长上下文输入和输出 token 生产效率为核心管理对象。",
                "目标是把模型推理从单点压测转成可规模化运营的 token 产能体系。",
            ],
            [
                "请求规模",
                f"历史样本累计 {request_scale} 请求，成功率 100%，可作为本轮容量测算的业务底座。",
                "先按已观测工服请求规模做产能口径，后续随业务峰值补充并发水位。",
            ],
            [
                "Token 规模",
                f"历史样本累计约 {token_scale} token，输入 token 占比约 {input_ratio}。",
                "成本和容量压力主要来自输入侧长上下文，优化重点放在 prefill、KV 和缓存复用。",
            ],
            [
                "长上下文压力",
                f"32K 以上请求占比约 {long_context_ratio}，P95 总长度已进入 100K 级别。",
                "适用大参数+长上下文场景，需要同时关注 TTFT、TPOT、吞吐和显存占用。",
            ],
            [
                "指标口径",
                "请求量、token 量、输入占比和 32K 以上占比来自 context analysis 历史数据。",
                "已有结果标注为已有结论/历史数据；缺少线上复测的数据标注待补测/待资源验证。",
            ],
        ],
    )
    _write_section_title(ws, row + 1, "本阶段目标")
    _append_table(
        ws,
        row + 2,
        ["目标", "验收关注", "状态"],
        [
            ["模型量化", "用 H200 BF16/FP8 历史数据量化吞吐和 TPOT 收益。", "已有结论/历史数据"],
            ["PD 分离资源方案", "形成 H200 做 PD、H200 做 P、H20 做 D 的资源验证路径。", "待资源验证"],
            ["汇报材料", "把背景、量化结论、资源方案和数据来源写入 workbook。", "本任务补充"],
        ],
    )


def _add_context_distribution_chart(ws, context: dict[str, object]) -> None:
    buckets = context.get("input_buckets", [])
    if not isinstance(buckets, list):
        buckets = []
    long_buckets = [
        bucket
        for bucket in buckets
        if isinstance(bucket, dict) and int(bucket.get("sort_key", 0) or 0) >= 32_768
    ]
    if not long_buckets:
        ws["J6"] = "图表待补测：缺少 32K 以上上下文分桶数据"
        ws["J6"].font = Font(bold=True, color="9C6500")
        ws["J6"].fill = PatternFill("solid", fgColor="FFF2CC")
        return

    start_row = 6
    start_col = 10
    rows = [
        [
            str(bucket["range_label"]),
            int(bucket["request_count"]),
            float(bucket["request_pct"]),
        ]
        for bucket in long_buckets
    ]
    last_row = _write_chart_source_table(
        ws,
        start_row,
        start_col,
        ["输入长度分桶", "请求数", "请求占比%"],
        rows,
    )
    for col_offset, width in enumerate((16, 12, 14)):
        ws.column_dimensions[ws.cell(row=1, column=start_col + col_offset).column_letter].width = width
    for row_idx in range(start_row + 1, last_row + 1):
        ws.cell(row=row_idx, column=start_col + 2).number_format = "0.0"

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "32K 以上长上下文输入长度分布（请求占比）"
    chart.y_axis.title = "请求占比（%）"
    chart.x_axis.title = "输入长度分桶"
    chart.height = 7
    chart.width = 11
    data = Reference(ws, min_col=start_col + 2, min_row=start_row, max_row=last_row)
    categories = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "J12")


def _add_fp8_comparison_chart(ws, fp8_summary: dict[str, dict[str, float | int | str]]) -> None:
    rows = []
    for model_name in ("32B", "72B"):
        summary = fp8_summary.get(model_name)
        if not _has_fp8_result(summary):
            continue
        rows.append(
            [
                model_name,
                float(summary["avg_bf16_throughput"]),
                float(summary["avg_fp8_throughput"]),
                float(summary["avg_throughput_gain_pct"]) / 100.0,
            ]
        )

    start_row = 6
    start_col = 10
    if not rows:
        ws.cell(row=start_row, column=start_col, value="图表待补测：缺少有效 FP8 对比数据")
        ws.cell(row=start_row, column=start_col).font = Font(bold=True, color="9C6500")
        ws.cell(row=start_row, column=start_col).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.column_dimensions["J"].width = 34
        return

    last_row = _write_chart_source_table(
        ws,
        start_row,
        start_col,
        ["模型", "BF16 吞吐", "FP8 吞吐", "吞吐提升"],
        rows,
    )
    for col_offset, width in enumerate((10, 14, 14, 12)):
        ws.column_dimensions[ws.cell(row=1, column=start_col + col_offset).column_letter].width = width
    for row_idx in range(start_row + 1, last_row + 1):
        ws.cell(row=row_idx, column=start_col + 1).number_format = "0.0"
        ws.cell(row=row_idx, column=start_col + 2).number_format = "0.0"
        ws.cell(row=row_idx, column=start_col + 3).number_format = "0.0%"

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "H200 FP8 vs BF16 吞吐对比"
    chart.y_axis.title = "输出 tokens 总吞吐"
    chart.x_axis.title = "模型"
    chart.height = 7
    chart.width = 12
    data = Reference(ws, min_col=start_col + 1, max_col=start_col + 2, min_row=start_row, max_row=last_row)
    categories = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "J11")


def _has_fp8_result(summary: dict[str, float | int | str] | None) -> bool:
    return bool(summary and summary.get("status") == "已有结论/历史数据")


def _write_model_quantization(ws, fp8_summary: dict[str, dict[str, float | int | str]]) -> None:
    start_row = ws.max_row + 2
    row = _write_section_title(ws, start_row, "模型量化结论")
    table_rows = []
    for model_name in ("32B", "72B"):
        summary = fp8_summary.get(model_name)
        if not _has_fp8_result(summary):
            reason = summary.get("reason", "缺少数据") if summary else "缺少数据"
            matched_rows = str(summary.get("matched_rows", 0)) if summary else "0"
            table_rows.append([model_name, "待补测", matched_rows, "待补测", "待补测", reason])
            continue
        throughput_gain = _format_pct(summary["avg_throughput_gain_pct"])
        tpot_ratio = _format_pct(summary["avg_tpot_ratio_pct"])
        table_rows.append(
            [
                model_name,
                str(summary["status"]),
                str(summary["matched_rows"]),
                throughput_gain,
                tpot_ratio,
                "已有结论/历史数据：FP8 相比 BF16 吞吐提升明显，TPOT 降低，适合作为 GLM5.1 或 GLM5.2 FP8 部署口径。",
            ]
        )
    row = _append_table(
        ws,
        row,
        ["模型", "状态", "匹配样本数", "吞吐约提升", "TPOT 约为 BF16", "备注"],
        table_rows,
    )
    available_models = [model_name for model_name in ("32B", "72B") if _has_fp8_result(fp8_summary.get(model_name))]
    if available_models:
        model_list = "、".join(available_models)
        gain_text = "、".join(
            f"{model_name} {_format_pct(fp8_summary[model_name]['avg_throughput_gain_pct'])}"
            for model_name in available_models
        )
        summary_row = [
            "优先采用 FP8 推理口径",
            f"{model_list} 的 H200 历史配对数据已可用，吞吐约提升 {gain_text}，可进入领导汇报的量化收益页。",
            "线上峰值、混部和更长上下文组合仍需待资源验证。",
        ]
    else:
        summary_row = [
            "量化收益待补测",
            "当前 H200 BF16/FP8 可比数据缺失，量化收益页先标注待补测/缺少数据。",
            "补齐 BF16 与 FP8 同输入长度、输出长度、并发数的配对样本后再写收益结论。",
        ]

    detail_rows = [summary_row]
    if _has_fp8_result(fp8_summary.get("72B")):
        detail_rows.append(
            [
                "72B 作为大参数重点展示",
                f"72B FP8 已有 {fp8_summary['72B']['matched_rows']} 条匹配样本，可支撑大参数模型容量申请的历史数据依据。",
                "GLM5.1/GLM5.2 实际模型版本落地后补测 TTFT、TPOT 和显存水位。",
            ]
        )
    else:
        reason = fp8_summary.get("72B", {}).get("reason", "缺少数据")
        detail_rows.append(
            [
                "72B 待补测",
                f"72B 当前 {reason}，暂不写大参数量化收益结论。",
                "补齐 72B BF16/FP8 配对数据后再更新领导汇报口径。",
            ]
        )

    _write_section_title(ws, row + 1, "汇报建议")
    _append_table(
        ws,
        row + 2,
        ["结论", "说明", "风险/下一步"],
        detail_rows,
    )


def _write_pd_plan(ws) -> None:
    start_row = ws.max_row + 2
    row = _write_section_title(ws, start_row, "PD 分离资源方案")
    row = _append_table(
        ws,
        row,
        ["模块", "方案", "汇报口径"],
        [
            [
                "原理",
                "Prefill 负责长上下文首轮计算，Decode 负责逐 token 生成；两阶段拆开后可按资源特性分别扩容。",
                "适用大参数+长上下文，目标是降低 TTFT、稳定 TPOT，并提升整体 token 工厂吞吐。",
            ],
            [
                "基线",
                "H200 做 PD：同一类 H200 资源同时承担 Prefill 和 Decode。",
                "作为当前验证起点，先建立端到端延迟、吞吐和显存基线。",
            ],
            [
                "同构分离",
                "H200 做 P、H200 做 D：Prefill 与 Decode 独立池化，但都使用 H200。",
                "用于验证 PD 调度收益，排除异构硬件差异对结果的影响。",
            ],
            [
                "异构方案",
                "H200 做 P、H20 做 D：用 H200 承担长上下文 Prefill，用 H20 承担 Decode。",
                "异构 H200+H20 方案可把高带宽高算力 H200 留给 prefill，H20 承接生成侧容量。",
            ],
        ],
    )
    row = _write_section_title(ws, row + 1, "资源申请与验证计划")
    _append_table(
        ws,
        row,
        ["事项", "要求", "状态"],
        [
            [
                "资源申请口径",
                "以 GLM5.1 或 GLM5.2 FP8 部署为口径申请资源，按 72B 大参数长上下文优先验证。",
                "当前待申请资源",
            ],
            [
                "带宽/互联要求",
                "P/D 之间需要稳定低延迟互联和足够带宽传输 KV 或中间状态；跨机部署需重点压测链路抖动。",
                "待资源验证",
            ],
            [
                "验证指标",
                "记录 TTFT、TPOT、输出 tokens 总吞吐、GPU 显存水位、跨节点带宽占用和失败率。",
                "待补测",
            ],
            [
                "验证顺序",
                "先跑 H200 做 PD 基线，再跑 H200 做 P、H200 做 D，最后跑 H200 做 P、H20 做 D 的异构 H200+H20 方案。",
                "待资源验证",
            ],
        ],
    )


def _write_appendix(ws) -> None:
    start_row = ws.max_row + 2
    row = _write_section_title(ws, start_row, "数据来源与状态")
    _append_table(
        ws,
        row,
        ["数据项", "路径", "用途", "状态"],
        [
            [
                "上下文总览",
                "outputs/context_analysis_20260609_034248/01_overview.json",
                "请求规模、token 规模、输入占比、延迟概览。",
                "已有结论/历史数据",
            ],
            [
                "输入长度分桶",
                "outputs/context_analysis_20260609_034248/02_input_buckets.csv",
                "32K 以上长上下文占比。",
                "已有结论/历史数据",
            ],
            [
                "H200 32B BF16",
                "data/H200/32B/1.csv",
                "32B FP8 对比基线。",
                "已有结论/历史数据",
            ],
            [
                "H200 32B FP8",
                "data/H200/32B-FP8/1.csv",
                "32B FP8 吞吐和 TPOT 对比。",
                "已有结论/历史数据",
            ],
            [
                "H200 72B BF16",
                "data/H200/72B/2.csv",
                "72B FP8 对比基线。",
                "已有结论/历史数据",
            ],
            [
                "H200 72B FP8",
                "data/H200/72B-FP8/2.csv",
                "72B FP8 吞吐和 TPOT 对比。",
                "已有结论/历史数据",
            ],
            [
                "PD 分离验证",
                "待资源验证",
                "H200 做 PD、H200 做 P、H20 做 D 以及异构 H200+H20 实测。",
                "待申请资源/待补测",
            ],
        ],
    )


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_context_summary(repo_root: Path) -> dict[str, object]:
    out_dir = repo_root / "outputs" / "context_analysis_20260609_034248"
    overview = json.loads((out_dir / "01_overview.json").read_text(encoding="utf-8"))
    input_rows = _read_csv_rows(out_dir / "02_input_buckets.csv")
    total_requests = int(float(overview.get("total_requests", 0) or 0))
    total_tokens = float(overview.get("total_tokens", 0) or 0)
    input_tokens = float(overview.get("total_input_tokens", 0) or 0)
    long_requests = sum(
        int(row["request_count"])
        for row in input_rows
        if row["range_label"] in {"32K-64K", "64K-128K", "128K+"}
    )
    input_buckets = []
    for row in input_rows:
        request_count = int(row["request_count"])
        input_buckets.append(
            {
                "range_label": row["range_label"],
                "sort_key": int(float(row["sort_key"])),
                "request_count": request_count,
                "request_pct": request_count / total_requests * 100.0 if total_requests else 0.0,
            }
        )
    return {
        "total_requests": total_requests,
        "total_tokens_billion": total_tokens / 1_000_000_000,
        "input_token_ratio": input_tokens / total_tokens if total_tokens else 0,
        "long_context_ratio": long_requests / total_requests if total_requests else 0,
        "input_buckets": input_buckets,
    }


def _load_h200_csv(path: Path) -> dict[tuple[float, float, int], dict[str, float]]:
    rows = {}
    for row in _read_csv_rows(path):
        key = (
            float(row["输入长度"]),
            float(row["输出长度"]),
            int(float(row["并发数"])),
        )
        rows[key] = {
            "throughput": float(row["输出tokens总吞吐"]),
            "avg_tpot_ms": float(row["平均增量时延（ms）"]),
            "avg_ttft_ms": float(row["平均首tokens时延（ms）"]),
        }
    return rows


def _fp8_pending(reason: str) -> dict[str, float | int | str]:
    return {
        "status": "待补测",
        "reason": reason,
        "matched_rows": 0,
    }


def _compare_h200_pair(base_path: Path, fp8_path: Path) -> dict[str, float | int | str]:
    missing_paths = [path for path in (base_path, fp8_path) if not path.exists()]
    if missing_paths:
        missing_text = "、".join(path.as_posix() for path in missing_paths)
        return _fp8_pending(f"缺少数据：{missing_text}")

    base = _load_h200_csv(base_path)
    fp8 = _load_h200_csv(fp8_path)
    keys = sorted(set(base) & set(fp8))
    valid_keys = [
        key
        for key in keys
        if base[key]["throughput"] > 0
        and base[key]["avg_tpot_ms"] > 0
        and fp8[key]["throughput"] > 0
        and fp8[key]["avg_tpot_ms"] > 0
    ]
    if not valid_keys:
        return _fp8_pending("缺少数据：BF16/FP8 无可比样本")

    throughput_ratios = [fp8[k]["throughput"] / base[k]["throughput"] for k in valid_keys]
    tpot_ratios = [fp8[k]["avg_tpot_ms"] / base[k]["avg_tpot_ms"] for k in valid_keys]
    return {
        "status": "已有结论/历史数据",
        "reason": "已匹配 H200 BF16/FP8 历史数据",
        "matched_rows": len(valid_keys),
        "avg_bf16_throughput": mean(base[k]["throughput"] for k in valid_keys),
        "avg_fp8_throughput": mean(fp8[k]["throughput"] for k in valid_keys),
        "avg_throughput_gain_pct": (mean(throughput_ratios) - 1.0) * 100.0,
        "avg_tpot_ratio_pct": mean(tpot_ratios) * 100.0,
    }


def compute_fp8_summary(repo_root: Path) -> dict[str, dict[str, float | int | str]]:
    h200 = repo_root / "data" / "H200"
    return {
        "32B": _compare_h200_pair(h200 / "32B" / "1.csv", h200 / "32B-FP8" / "1.csv"),
        "72B": _compare_h200_pair(h200 / "72B" / "2.csv", h200 / "72B-FP8" / "2.csv"),
    }


def build_workbook(repo_root: Path) -> Workbook:
    context = load_context_summary(repo_root)
    fp8_summary = compute_fp8_summary(repo_root)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for spec in SHEET_SPECS:
        ws = wb.create_sheet(spec["title"])
        _write_header(ws, spec)
        if spec["title"] in PRINCIPLE_DIAGRAMS:
            _write_principle_diagram(ws, PRINCIPLE_DIAGRAMS[spec["title"]])
        if spec["title"] == "00_目录与版本":
            _write_sheet_index(ws)
        if spec["title"] == "02_总览结论":
            _write_summary(ws)
        if spec["title"] == "01_背景与目标":
            _write_background(ws, context)
            _add_context_distribution_chart(ws, context)
        if spec["title"] == "04_模型量化":
            _write_model_quantization(ws, fp8_summary)
            _add_fp8_comparison_chart(ws, fp8_summary)
        if spec["title"] == "08_PD分离":
            _write_pd_plan(ws)
        if spec["title"] == "12_数据附录":
            _write_appendix(ws)
    return wb


def default_repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def resolve_output_path(repo_root: Path, output: Path | str | None) -> Path:
    output_path = Path("inference-report") / OUTPUT_FILENAME if output is None else Path(output)
    if output_path.is_absolute():
        return output_path
    return repo_root / output_path


def save_report(repo_root: Path, output: Path | str | None = None) -> Path:
    output_path = resolve_output_path(repo_root, output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(repo_root)
    workbook.save(output_path)
    return output_path


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help=f"输出 xlsx 路径；相对路径按 repo root 解析。默认 inference-report/{OUTPUT_FILENAME}",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    output_path = save_report(default_repo_root(), args.output)
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
