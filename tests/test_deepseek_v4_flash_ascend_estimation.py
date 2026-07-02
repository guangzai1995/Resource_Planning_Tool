from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from scripts.estimate_deepseek_v4_flash_ascend import (
    DEFAULT_BASE_MODEL_ASSUMPTION,
    DEFAULT_GPU_ASSUMPTIONS,
    DEFAULT_MODEL_ASSUMPTION,
    DEFAULT_RUNTIME_SCENARIOS,
    DEFAULT_TARGET_MODEL_ASSUMPTION,
    SOURCE_COLUMNS,
    BaselineRow,
    GpuAssumption,
    build_estimates,
    build_workbook,
    estimate_row,
    load_p800_baseline,
    main,
    public_reference_dataframe,
    required_memory_per_card_gb,
    summarize_910b_reference,
)


def test_required_memory_per_card_changes_with_weight_precision():
    bf16 = required_memory_per_card_gb(
        model=DEFAULT_MODEL_ASSUMPTION.with_updates(weight_bytes=2.0),
        gpu_count=8,
        input_tokens=512,
        output_tokens=1024,
        concurrency=5,
        framework_overhead_gb=3.0,
    )
    int8 = required_memory_per_card_gb(
        model=DEFAULT_MODEL_ASSUMPTION.with_updates(weight_bytes=1.0),
        gpu_count=8,
        input_tokens=512,
        output_tokens=1024,
        concurrency=5,
        framework_overhead_gb=3.0,
    )
    assert bf16 > int8
    assert int8 > 3.0


def test_compute_scale_uses_injected_target_hardware_ratio():
    baseline = BaselineRow(
        excel_row=2,
        input_tokens=512,
        output_tokens=1024,
        concurrency=5,
        throughput_tokens_s=100.0,
        ttft_p90_ms=400.0,
        ttft_p99_ms=500.0,
        ttft_max_ms=550.0,
        ttft_mean_ms=300.0,
        decode_latency_p90_ms=40.0,
        decode_latency_p99_ms=45.0,
        decode_latency_max_ms=60.0,
        decode_latency_mean_ms=35.0,
    )
    target_gpu = GpuAssumption("TEST-ACCEL", 128.0, 4000.0, 780.0, "test_assumption")
    neutral_model = DEFAULT_MODEL_ASSUMPTION.with_updates(weight_bytes=1.0)
    result = estimate_row(
        baseline=baseline,
        target_gpu=target_gpu,
        base_gpu=DEFAULT_GPU_ASSUMPTIONS["P800"],
        base_model=neutral_model,
        target_model=neutral_model,
        runtime=DEFAULT_RUNTIME_SCENARIOS["base"],
    )
    assert result["target_gpu"] == "TEST-ACCEL"
    assert result["gpu_count"] == 8
    assert result["compute_scale"] == pytest.approx(780 / 280, rel=1e-6)
    assert result["decode_scale"] == pytest.approx(
        min(result["compute_scale"], result["bandwidth_scale"]),
        rel=1e-6,
    )
    assert result["输出tokens总吞吐"] == pytest.approx(100.0 * result["decode_scale"])
    assert 0.0 <= result["confidence"] <= 1.0


def test_default_model_and_p800_assumptions_match_confirmed_inputs():
    assert DEFAULT_GPU_ASSUMPTIONS["P800"].memory_gb == 96.0
    assert DEFAULT_BASE_MODEL_ASSUMPTION.name == "DeepSeek R1 INT8 baseline proxy"
    assert DEFAULT_BASE_MODEL_ASSUMPTION.total_params_b == 671.0
    assert DEFAULT_BASE_MODEL_ASSUMPTION.active_params_b == 37.0
    assert DEFAULT_BASE_MODEL_ASSUMPTION.weight_bytes == 1.0
    assert DEFAULT_TARGET_MODEL_ASSUMPTION.name == "DeepSeek V4 Flash INT8 public proxy"
    assert DEFAULT_TARGET_MODEL_ASSUMPTION.total_params_b == 284.0
    assert DEFAULT_TARGET_MODEL_ASSUMPTION.active_params_b == 13.0
    assert DEFAULT_TARGET_MODEL_ASSUMPTION.weight_bytes == 1.0


def test_model_difference_changes_compute_and_prefill_scales():
    baseline = BaselineRow(
        excel_row=2,
        input_tokens=512,
        output_tokens=1024,
        concurrency=5,
        throughput_tokens_s=100.0,
        ttft_p90_ms=400.0,
        ttft_p99_ms=500.0,
        ttft_max_ms=550.0,
        ttft_mean_ms=300.0,
        decode_latency_p90_ms=40.0,
        decode_latency_p99_ms=45.0,
        decode_latency_max_ms=60.0,
        decode_latency_mean_ms=35.0,
    )
    base_model = DEFAULT_MODEL_ASSUMPTION.with_updates(active_params_b=37.0)
    target_model = DEFAULT_MODEL_ASSUMPTION.with_updates(active_params_b=74.0)
    result = estimate_row(
        baseline=baseline,
        target_gpu=DEFAULT_GPU_ASSUMPTIONS["P800"],
        base_gpu=DEFAULT_GPU_ASSUMPTIONS["P800"],
        base_model=base_model,
        target_model=target_model,
        runtime=DEFAULT_RUNTIME_SCENARIOS["base"],
    )
    assert result["compute_scale"] == pytest.approx(0.5, rel=1e-4)
    assert result["输出tokens总吞吐"] < baseline.throughput_tokens_s
    assert result["平均首tokens时延（ms）"] > baseline.ttft_mean_ms


def write_baseline_workbook(path: Path) -> None:
    df = pd.DataFrame(
        [
            {
                "输入长度": 512,
                "输出长度": 1024,
                "并发数": 5,
                "输出tokens总吞吐": 117.0,
                "首tokens时延TP90（ms）": 476.0,
                "首tokens时延TP99（ms）": 500.0,
                "最大首tokens时延（ms）": 501.0,
                "平均首tokens时延（ms）": 381.0,
                "增量时延TP90（ms）": 43.0,
                "增量时延TP99（ms）": 44.0,
                "最大增量时延（ms）": 46.0,
                "平均增量时延（ms）": 42.0,
            },
            {
                "输入长度": 1024,
                "输出长度": 1024,
                "并发数": 10,
                "输出tokens总吞吐": 150.0,
                "首tokens时延TP90（ms）": 600.0,
                "首tokens时延TP99（ms）": 650.0,
                "最大首tokens时延（ms）": 700.0,
                "平均首tokens时延（ms）": 500.0,
                "增量时延TP90（ms）": 55.0,
                "增量时延TP99（ms）": 58.0,
                "最大增量时延（ms）": 70.0,
                "平均增量时延（ms）": 50.0,
            },
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="671B-P800-8测试数据", index=False)


def test_load_p800_baseline_preserves_excel_row_numbers(tmp_path):
    workbook = tmp_path / "资源规划工具.xlsx"
    write_baseline_workbook(workbook)
    rows = load_p800_baseline(workbook)
    assert len(rows) == 2
    assert rows[0].excel_row == 2
    assert rows[1].excel_row == 3
    assert rows[0].throughput_tokens_s == 117.0


def test_build_estimates_creates_910b_and_910c_rows(tmp_path):
    workbook = tmp_path / "资源规划工具.xlsx"
    write_baseline_workbook(workbook)
    baseline_rows = load_p800_baseline(workbook)
    estimates = build_estimates(baseline_rows)
    assert set(estimates["target_gpu"]) == {"910B", "910C"}
    assert len(estimates) == 4
    assert list(estimates.columns[: len(SOURCE_COLUMNS)]) == SOURCE_COLUMNS
    assert {"base_row", "compute_scale", "bandwidth_scale", "decode_scale", "confidence"}.issubset(estimates.columns)


def write_reference_csv(path: Path, throughput: float) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(
            [
                "输入长度,输出长度,并发数,输出tokens总吞吐,首tokens时延TP90（ms）,首tokens时延TP99（ms）,最大首tokens时延（ms）,平均首tokens时延（ms）,增量时延TP90（ms）,增量时延TP99（ms）,最大增量时延（ms）,平均增量时延（ms）",
                f"2048,1200,1,{throughput},320,330,340,310,30,35,55,29",
                f"2048,1200,4,{throughput * 2},1100,1110,1120,1000,32,38,58,30",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def test_910b_reference_summary_does_not_change_estimates(tmp_path):
    workbook = tmp_path / "资源规划工具.xlsx"
    write_baseline_workbook(workbook)
    baseline_rows = load_p800_baseline(workbook)
    before = build_estimates(baseline_rows)

    write_reference_csv(tmp_path / "data" / "910B1" / "72B-AWQ" / "8.csv", 10.0)
    first_summary = summarize_910b_reference(tmp_path / "data")
    write_reference_csv(tmp_path / "data" / "910B1" / "72B-AWQ" / "8.csv", 9999.0)
    second_summary = summarize_910b_reference(tmp_path / "data")
    after = build_estimates(baseline_rows)

    assert first_summary.loc[0, "mean_throughput_tokens_s"] != second_summary.loc[0, "mean_throughput_tokens_s"]
    pd.testing.assert_frame_equal(before, after)


def test_910b_reference_summary_reads_uppercase_csv(tmp_path):
    write_reference_csv(tmp_path / "data" / "910B3" / "72B-AWQ" / "8.CSV", 12.0)
    summary = summarize_910b_reference(tmp_path / "data")
    assert len(summary) == 1
    assert summary.loc[0, "source_file"].endswith("8.CSV")


def test_public_hardware_reference_drives_default_910c_estimates(tmp_path):
    workbook = tmp_path / "资源规划工具.xlsx"
    write_baseline_workbook(workbook)
    public_refs = public_reference_dataframe()
    assert public_refs["metric_summary"].str.contains("780 BF16 TFLOPS").any()

    estimates = build_estimates(load_p800_baseline(workbook))
    first_910c = estimates[estimates["target_gpu"] == "910C"].iloc[0]
    expected_hardware_scale = 780 / 280
    expected_model_scale = (
        DEFAULT_BASE_MODEL_ASSUMPTION.active_params_b
        / DEFAULT_TARGET_MODEL_ASSUMPTION.active_params_b
    )
    assert first_910c["compute_scale"] == pytest.approx(
        expected_hardware_scale * expected_model_scale,
        rel=1e-6,
    )
    assert "public_cloudmatrix384" in first_910c["assumption_notes"]
    assert "DeepSeek-R1/P800 INT8 baseline proxy -> DeepSeek-V4-Flash public 284B/13B" in first_910c[
        "assumption_notes"
    ]


def test_build_workbook_writes_expected_sheets(tmp_path):
    workbook = tmp_path / "资源规划工具.xlsx"
    output = tmp_path / "estimate.xlsx"
    write_baseline_workbook(workbook)
    write_reference_csv(tmp_path / "data" / "910B3" / "72B-AWQ" / "8.csv", 20.0)

    build_workbook(
        source_workbook=workbook,
        data_dir=tmp_path / "data",
        output_path=output,
    )

    loaded = load_workbook(output, data_only=False)
    assert loaded.sheetnames == [
        "00_估算说明",
        "01_假设表",
        "02_910B_8卡估算",
        "03_910C_8卡估算",
        "04_P800基线",
        "05_910B低利用率参考",
        "06_联网参考数据",
    ]
    assert loaded["00_估算说明"]["A1"].value == "DeepSeek v4 flash 910B/910C 8卡估算"
    assert loaded["00_估算说明"]["B8"].value.startswith("DeepSeek V4-Flash")
    assert loaded["02_910B_8卡估算"]["A1"].value == "输入长度"
    assert loaded["03_910C_8卡估算"]["A1"].value == "输入长度"
    assert loaded["05_910B低利用率参考"]["J2"].value == "低利用率参考，不参与估算校准"
    assert loaded["06_联网参考数据"]["A1"].value == "source_title"
    assert "CloudMatrix384" in loaded["06_联网参考数据"]["A4"].value


def test_main_generates_output_file(tmp_path):
    workbook = tmp_path / "资源规划工具.xlsx"
    output = tmp_path / "estimate.xlsx"
    write_baseline_workbook(workbook)

    exit_code = main(
        [
            "--source-workbook",
            str(workbook),
            "--data-dir",
            str(tmp_path / "data"),
            "--output",
            str(output),
        ]
    )

    assert exit_code == 0
    assert output.exists()
