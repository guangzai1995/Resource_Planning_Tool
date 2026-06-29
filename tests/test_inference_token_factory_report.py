import hashlib
import json
from pathlib import Path
import time

from openpyxl import load_workbook
import pytest

from scripts.build_inference_token_factory_report import (
    OUTPUT_FILENAME,
    REPORT_VERSION,
    SHEET_SPECS,
    STATUS_LABELS,
    build_workbook,
    compute_fp8_summary,
    load_context_summary,
    main,
    resolve_output_path,
    save_report,
)


EXPECTED_SHEETS = [
    "00_目录与版本",
    "01_背景与目标",
    "02_总览结论",
    "03_技术路线地图",
    "04_模型量化",
    "05_KVCache量化",
    "06_Prefix_KV命中",
    "07_投机解码",
    "08_PD分离",
    "09_MOE专家并行",
    "10_连续批处理",
    "11_汇总建议与资源计划",
    "12_数据附录",
]

PRINCIPLE_EXPECTATIONS = {
    "04_模型量化": ["BF16/FP16 权重与激活", "FP8/量化模型", "显存下降", "吞吐提升"],
    "05_KVCache量化": ["KV Cache Block", "fp16/bf16 KV", "fp8 KV", "fp4 H200 不支持"],
    "06_Prefix_KV命中": ["共享前缀", "Prefix Hash", "复用 KV", "TTFT 降低"],
    "07_投机解码": ["Draft 候选 token", "Verify", "Accept/Reject", "TPOT 降低"],
    "08_PD分离": ["Prefill(H200)", "KV 传输", "Decode(H200/H20)", "H200 P + H20 D"],
    "09_MOE专家并行": ["Router", "多 GPU Experts", "All2All 聚合", "GLM5.1/GLM5.2 FP8"],
    "10_连续批处理": ["Scheduler", "新请求 Prefill", "旧请求 Decode", "框架原生支持"],
}


def assert_cell_has_table_style(cell):
    assert cell.alignment.wrap_text is True
    assert cell.border.left.style == "thin"


def assert_cell_has_table_header_style(cell):
    assert cell.fill.fgColor.rgb in {"00D9EAF7", "D9EAF7"}
    assert cell.font.bold is True


def assert_cell_has_principle_node_style(cell):
    assert cell.fill.fgColor.rgb in {"00D9EAF7", "D9EAF7"}
    assert cell.font.bold is True
    assert cell.alignment.wrap_text is True
    assert cell.border.left.style == "thin"


def sheet_text(ws):
    return "\n".join(
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )


TOPIC_CONTENT_EXPECTATIONS = {
    "05_KVCache量化": ["显存节省约 50%", "fp4 因 H200 不支持", "TPS/Tokens/s"],
    "06_Prefix_KV命中": ["后续补测", "prefix_ratio", "cache hit rate"],
    "07_投机解码": ["候选 token 1/3/5", "接受率", "TPOT"],
    "09_MOE专家并行": ["需要验证测试", "GLM5.1 或 GLM5.2 FP8", "All2All"],
    "10_连续批处理": ["原生支持", "无法关闭", "无需独立测试"],
    "11_汇总建议与资源计划": [
        "FP8/量化优先落地",
        "Prefix/KV 命中补系统数据",
        "PD 分离申请异构资源验证",
        "MOE EP 按 GLM5.1/GLM5.2 FP8 申请资源验证",
        "Continuous Batching 已具备",
    ],
}


def assert_topic_content(wb):
    for sheet_name, expected_keywords in TOPIC_CONTENT_EXPECTATIONS.items():
        text = sheet_text(wb[sheet_name])
        for keyword in expected_keywords:
            assert keyword in text


def file_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write_context_files(
    repo_root: Path,
    *,
    total_requests: int = 100,
    total_tokens: int = 1000,
    total_input_tokens: int = 900,
    long_requests: int = 50,
):
    out_dir = repo_root / "outputs" / "context_analysis_20260609_034248"
    out_dir.mkdir(parents=True)
    (out_dir / "01_overview.json").write_text(
        json.dumps(
            {
                "total_requests": str(total_requests),
                "total_tokens": str(total_tokens),
                "total_input_tokens": str(total_input_tokens),
            }
        ),
        encoding="utf-8",
    )
    (out_dir / "02_input_buckets.csv").write_text(
        "\n".join(
            [
                "range_label,sort_key,request_count,pct,total_in_all",
                f"0-512,1,{max(total_requests - long_requests, 0)},0,0",
                f"32K-64K,32768,{long_requests},0,0",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def write_h200_csv(
    path: Path,
    throughput: float,
    tpot_ms: float,
    *,
    input_len: int = 512,
    output_len: int = 1024,
    concurrency: int = 1,
):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(
            [
                "输入长度,输出长度,并发数,输出tokens总吞吐,首tokens时延TP90（ms）,首tokens时延TP99（ms）,最大首tokens时延（ms）,平均首tokens时延（ms）,增量时延TP90（ms）,增量时延TP99（ms）,最大增量时延（ms）,平均增量时延（ms）",
                f"{input_len},{output_len},{concurrency},{throughput},1,1,1,10,1,1,1,{tpot_ms}",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def test_static_sheet_specs_and_status_labels():
    assert REPORT_VERSION == "v2.0"
    assert [item["title"] for item in SHEET_SPECS] == EXPECTED_SHEETS
    assert STATUS_LABELS == [
        "已验证",
        "有结论但缺系统数据",
        "待验证",
        "已具备无需独立测试",
    ]


def test_build_workbook_creates_expected_sheets():
    wb = build_workbook(Path("."))
    assert wb.sheetnames == EXPECTED_SHEETS


def test_workbook_contains_required_sheet_headers(tmp_path):
    output = tmp_path / "report.xlsx"
    wb = build_workbook(Path("."))
    wb.save(output)
    loaded = load_workbook(output)
    assert loaded.sheetnames == EXPECTED_SHEETS
    for sheet_name in EXPECTED_SHEETS:
        ws = loaded[sheet_name]
        assert ws["A1"].value == "推理 Token 工厂汇报"
        assert ws["B2"].value == REPORT_VERSION
        assert ws["A4"].value == "当前状态"

    index_ws = loaded["00_目录与版本"]
    assert index_ws["A7"].value == "Sheet"
    assert_cell_has_table_header_style(index_ws["A7"])
    assert index_ws.cell(row=index_ws.max_row, column=2).value == "12_数据附录"
    assert_cell_has_table_style(index_ws.cell(row=index_ws.max_row, column=2))


def test_summary_sheet_has_expected_technology_states(tmp_path):
    output = tmp_path / "report.xlsx"
    wb = build_workbook(Path("."))
    wb.save(output)
    ws = load_workbook(output)["02_总览结论"]
    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert "模型量化" in values
    assert "MOE 专家并行" in values
    assert "待验证" in values
    assert "连续批处理" in values
    assert "已具备无需独立测试" in values
    assert ws["A7"].value == "技术点"
    assert_cell_has_table_header_style(ws["A7"])
    assert ws.cell(row=ws.max_row, column=1).value == "连续批处理"
    assert_cell_has_table_style(ws.cell(row=ws.max_row, column=1))


def test_load_context_summary_from_existing_outputs():
    summary = load_context_summary(Path("."))
    assert summary["total_requests"] == 1_008_098
    assert summary["total_tokens_billion"] == pytest.approx(59.36, rel=0.01)
    assert summary["input_token_ratio"] == pytest.approx(0.989, rel=0.01)
    assert summary["long_context_ratio"] == pytest.approx(0.623, rel=0.02)


def test_compute_h200_fp8_summary_from_existing_csvs():
    summary = compute_fp8_summary(Path("."))
    assert summary["72B"]["matched_rows"] >= 30
    assert summary["72B"]["status"] == "已有结论/历史数据"
    assert summary["72B"]["avg_throughput_gain_pct"] == pytest.approx(32.4, rel=0.08)
    assert summary["72B"]["avg_tpot_ratio_pct"] == pytest.approx(77.5, rel=0.08)
    assert summary["32B"]["matched_rows"] >= 30
    assert summary["32B"]["status"] == "已有结论/历史数据"
    assert summary["32B"]["avg_throughput_gain_pct"] == pytest.approx(30.3, rel=0.08)
    assert summary["32B"]["avg_tpot_ratio_pct"] == pytest.approx(77.8, rel=0.08)


def test_each_technology_sheet_has_principle_diagram(tmp_path):
    output = tmp_path / "report.xlsx"
    wb = build_workbook(Path("."))
    wb.save(output)
    loaded = load_workbook(output)
    node_columns = ["A", "C", "E", "G"]
    arrow_columns = ["B", "D", "F"]
    for sheet_name, expected_labels in PRINCIPLE_EXPECTATIONS.items():
        ws = loaded[sheet_name]
        assert ws["A6"].value == "V2.2 方案：原理简图"

        for col, expected_label in zip(node_columns, expected_labels):
            cell = ws[f"{col}8"]
            assert cell.value == expected_label
            assert_cell_has_principle_node_style(cell)
            assert ws.column_dimensions[col].width >= 20

        for col in arrow_columns:
            cell = ws[f"{col}8"]
            assert cell.value == "→"
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.vertical == "center"
            assert cell.border.left.style == "thin"
            assert ws.column_dimensions[col].width == 6

        assert ws["A10"].value == "说明"
        assert "快速解释技术机制" in ws["B10"].value
        for cell in (ws["A10"], ws["B10"]):
            assert cell.alignment.wrap_text is True
            assert cell.border.left.style == "thin"


def test_report_contains_key_business_and_quantization_conclusions():
    wb = build_workbook(Path("."))

    background_text = sheet_text(wb["01_背景与目标"])
    assert "100.8 万" in background_text
    assert "593.6 亿" in background_text
    assert "32K 以上" in background_text

    quantization_text = sheet_text(wb["04_模型量化"])
    assert "72B" in quantization_text
    assert "吞吐约提升" in quantization_text
    assert "32.4%" in quantization_text

    pd_text = sheet_text(wb["08_PD分离"])
    assert "H200 做 P、H20 做 D" in pd_text

    appendix_text = sheet_text(wb["12_数据附录"])
    assert "data/H200/72B-FP8/2.csv" in appendix_text


def test_report_contains_required_topic_body_content():
    wb = build_workbook(Path("."))

    assert_topic_content(wb)


def test_report_contains_visual_charts():
    wb = build_workbook(Path("."))

    background_charts = wb["01_背景与目标"]._charts
    assert len(background_charts) >= 1
    assert any("上下文" in str(chart.title) for chart in background_charts)

    quantization_charts = wb["04_模型量化"]._charts
    assert len(quantization_charts) >= 1
    assert any(
        "FP8" in str(chart.title) or "吞吐" in str(chart.title)
        for chart in quantization_charts
    )


def assert_saved_report_content(output: Path):
    loaded = load_workbook(output)
    assert loaded.sheetnames == EXPECTED_SHEETS
    assert len(loaded["01_背景与目标"]._charts) >= 1
    assert len(loaded["04_模型量化"]._charts) >= 1

    background_text = sheet_text(loaded["01_背景与目标"])
    assert "100.8 万" in background_text

    pd_text = sheet_text(loaded["08_PD分离"])
    assert "H200 做 P、H20 做 D" in pd_text

    appendix_text = sheet_text(loaded["12_数据附录"])
    assert "data/H200/72B-FP8/2.csv" in appendix_text

    assert_topic_content(loaded)


def test_save_report_writes_reloadable_workbook_with_charts_and_key_content(tmp_path):
    output = tmp_path / "nested" / "report.xlsx"

    saved_path = save_report(Path("."), output)

    assert saved_path == output
    assert output.exists()
    assert_saved_report_content(output)


def test_save_report_is_byte_stable_across_repeated_generation(tmp_path):
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"

    save_report(Path("."), first)
    first_hash = file_sha256(first)
    time.sleep(2.1)
    save_report(Path("."), first)
    save_report(Path("."), second)

    assert file_sha256(first) == first_hash
    assert file_sha256(second) == first_hash
    assert_saved_report_content(first)
    assert_saved_report_content(second)


def test_cli_writes_requested_output_path_from_non_repo_cwd(tmp_path, monkeypatch, capsys):
    output = tmp_path / "cli" / "report.xlsx"
    monkeypatch.chdir(tmp_path)

    exit_code = main(["--output", str(output)])

    assert exit_code == 0
    assert output.exists()
    assert str(output) in capsys.readouterr().out
    assert_saved_report_content(output)


def test_resolve_output_path_uses_default_and_repo_root_for_relative_paths():
    repo_root = Path("/repo")

    assert resolve_output_path(repo_root, None) == repo_root / "inference-report" / OUTPUT_FILENAME
    assert resolve_output_path(repo_root, Path("custom/report.xlsx")) == repo_root / "custom" / "report.xlsx"
    assert resolve_output_path(repo_root, Path("/tmp/report.xlsx")) == Path("/tmp/report.xlsx")


def test_quantization_chart_degrades_when_all_fp8_data_is_missing(tmp_path):
    write_context_files(tmp_path)

    wb = build_workbook(tmp_path)
    ws = wb["04_模型量化"]

    assert ws._charts == []
    assert "图表待补测：缺少有效 FP8 对比数据" in sheet_text(ws)


def test_missing_h200_model_data_degrades_without_crashing(tmp_path):
    write_context_files(tmp_path)
    h200 = tmp_path / "data" / "H200"
    write_h200_csv(h200 / "32B" / "1.csv", throughput=100.0, tpot_ms=10.0)
    write_h200_csv(h200 / "32B-FP8" / "1.csv", throughput=130.0, tpot_ms=8.0)

    summary = compute_fp8_summary(tmp_path)
    assert summary["32B"]["status"] == "已有结论/历史数据"
    assert summary["32B"]["avg_throughput_gain_pct"] == pytest.approx(30.0)
    assert summary["72B"]["status"] == "待补测"
    assert "缺少数据" in summary["72B"]["reason"]

    wb = build_workbook(tmp_path)
    quantization_text = sheet_text(wb["04_模型量化"])
    assert "32B" in quantization_text
    assert "30.0%" in quantization_text
    assert "72B" in quantization_text
    assert "待补测" in quantization_text
    assert "缺少数据" in quantization_text
    assert "32B 的 H200 历史配对数据" in quantization_text
    assert "32B 和 72B 的 H200 历史配对数据均显示" not in quantization_text


def test_existing_h200_files_without_comparable_rows_are_marked_pending(tmp_path):
    write_context_files(tmp_path)
    h200 = tmp_path / "data" / "H200"
    write_h200_csv(h200 / "32B" / "1.csv", throughput=100.0, tpot_ms=10.0)
    write_h200_csv(h200 / "32B-FP8" / "1.csv", throughput=130.0, tpot_ms=8.0)
    write_h200_csv(h200 / "72B" / "2.csv", throughput=100.0, tpot_ms=10.0, input_len=512)
    write_h200_csv(h200 / "72B-FP8" / "2.csv", throughput=140.0, tpot_ms=7.0, input_len=2048)

    summary = compute_fp8_summary(tmp_path)
    assert summary["72B"]["status"] == "待补测"
    assert "缺少数据" in summary["72B"]["reason"]
    assert "无可比样本" in summary["72B"]["reason"]
    assert "avg_throughput_gain_pct" not in summary["72B"]

    wb = build_workbook(tmp_path)
    quantization_text = sheet_text(wb["04_模型量化"])
    assert "72B" in quantization_text
    assert "待补测" in quantization_text
    assert "无可比样本" in quantization_text
    assert "72B 40.0%" not in quantization_text
    assert "72B FP8 已有" not in quantization_text


def test_load_context_summary_handles_zero_totals(tmp_path):
    write_context_files(tmp_path, total_requests=0, total_tokens=0, total_input_tokens=0, long_requests=0)

    summary = load_context_summary(tmp_path)

    assert summary["total_requests"] == 0
    assert summary["total_tokens_billion"] == 0
    assert summary["input_token_ratio"] == 0
    assert summary["long_context_ratio"] == 0
