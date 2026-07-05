#!/usr/bin/env python3
"""
run_bench_multi.py — 多配置 vLLM OpenAI 接口批量性能基准测试

设计原则:
  复用 run_bench_serve.py 的全部 shim 注入逻辑（sys.modules 绕过 vllm/torch 依赖）
  以及 vllm_bench/serve.py 中已实现的 SSE 解析、指标计算（TTFT/TPOT/ITL/E2EL）
  和吞吐量统计。每组 (输入长度, 输出长度, 并发数) 配置调用一次 _serve.main_async()，
  收集返回的指标字典，最终汇总写入 CSV 和/或 XLSX。

参数映射:
  并发控制:  --parallel-nums N  →  _serve args.max_concurrency = N
                                    args.request_rate = inf（不限速，仅限并发）
  请求总数:  epochs × parallel_num  →  args.num_prompts
  长度控制:  --input-lens / --output-lens  →  args.input_len / args.output_len
             （serve.py 内部自动映射到 random_input_len / random_output_len）

最小依赖:
    pip install aiohttp numpy tqdm

可选（XLSX 导出）:
    pip install openpyxl

用法示例:
    python run_bench_multi.py \\
        --host 127.0.0.1 --port 8000 \\
        --model Qwen/Qwen2.5-7B-Instruct \\
        --input-lens 128 512 1024 2048 \\
        --output-lens 128 256 512 512 \\
        --parallel-nums 1 4 8 16 32 \\
        --epochs 3 \\
        --output-csv results/bench.csv \\
        --output-xlsx results/bench.xlsx
"""

import argparse
import asyncio
import copy
import csv
import hashlib
import importlib.util
import itertools
import json
import logging
import os
import sys
import time
from typing import Dict, List, Tuple

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S',
)
logger = logging.getLogger(__name__)

try:
    import openpyxl  # noqa: F401
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ─── 加载 run_bench_serve.py（触发全部 shim 注入，获取 _serve 模块）────────────
#
# run_bench_serve.py 在模块级完成了所有 sys.modules shim 注入，
# 并通过 importlib 按路径加载了本目录下 vllm_bench/serve.py（赋给 _serve）。
# 在此直接 exec 该文件，即可复用它的全部初始化逻辑，无需重复编写任何 shim。

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_RBS_PATH = os.path.join(_THIS_DIR, 'run_bench_serve.py')

if not os.path.isfile(_RBS_PATH):
    raise RuntimeError(
        f"找不到 run_bench_serve.py: {_RBS_PATH}\n"
        "run_bench_multi.py 需要与 run_bench_serve.py 放在同一目录。"
    )

_rbs_spec = importlib.util.spec_from_file_location('_run_bench_serve_shims', _RBS_PATH)
_rbs_mod = importlib.util.module_from_spec(_rbs_spec)
_rbs_spec.loader.exec_module(_rbs_mod)   # 执行所有 shim 注入 + 加载 vllm_bench/serve.py

# 获取已加载的 vllm_bench/serve.py 模块
_serve = _rbs_mod._serve

ASR_BACKEND = "openai-audio"
DEFAULT_ASR_DATASET_NAME = "custom_audio"


# ─── 构建 serve.py 所需的 argparse.Namespace ──────────────────────────────────

def _build_base_args(our_args: argparse.Namespace) -> argparse.Namespace:
    """
    使用 _serve.add_cli_args() 解析空参数列表，获取所有字段的默认值，
    然后将我们关心的字段覆盖进去。
    返回的 Namespace 可直接传给 _serve.main_async()。
    """
    # 用空列表解析，获取全部默认值（所有字段在 add_cli_args 中都有 default）
    inner_parser = argparse.ArgumentParser(add_help=False)
    _serve.add_cli_args(inner_parser)
    base = inner_parser.parse_args([])

    # ── 服务连接 ────────────────────────────────────────────────────────────
    base.model = our_args.model
    base.served_model_name = our_args.served_model_name or our_args.model
    base.backend = our_args.backend
    is_asr_backend = our_args.backend == ASR_BACKEND

    if our_args.base_url:
        # ── 完整 URL 模式（优先级高于 host:port）──────────────────────────────
        # 用法: --base-url https://example.com/openai/v1
        # 末尾不加斜杠，endpoint 只写路径部分（如 /completions）
        # 最终 api_url = base_url + endpoint
        # 例: https://example.com/openai/v1 + /completions
        #   → https://example.com/openai/v1/completions  ✓（以 completions 结尾）
        base.base_url = our_args.base_url.rstrip('/')
        # endpoint 不含版本前缀（base_url 已包含 /v1）
        if our_args.backend == 'openai-chat':
            base.endpoint = '/chat/completions'
        elif is_asr_backend:
            base.endpoint = '/audio/transcriptions'
        else:
            base.endpoint = '/completions'
        # HTTPS 时自动开启 SSL；如需跳过证书验证请使用 --insecure
        if our_args.insecure:
            base.insecure = True
    else:
        # ── host:port 模式 ────────────────────────────────────────────────────
        base.host = our_args.host
        base.port = our_args.port
        if our_args.backend == 'openai-chat':
            base.endpoint = '/v1/chat/completions'
        elif is_asr_backend:
            base.endpoint = '/v1/audio/transcriptions'
        else:
            base.endpoint = '/v1/completions'

    # API Key → serve.py 通过 header 传递
    if our_args.api_key:
        base.header = [f'Authorization=Bearer {our_args.api_key}']

    # ── tokenizer ────────────────────────────────────────────────────────────
    if is_asr_backend:
        base.skip_tokenizer_init = True
    elif our_args.tokenizer:
        base.tokenizer = our_args.tokenizer
    else:
        # 无 tokenizer 时跳过 tokenizer 初始化，使用近似字符串模式
        base.skip_tokenizer_init = True

    # ── 数据集 ──────────────────────────────────────────────────────────────
    if is_asr_backend:
        base.dataset_name = (
            getattr(our_args, "dataset_name", None) or DEFAULT_ASR_DATASET_NAME
        )
        base.dataset_path = getattr(our_args, "dataset_path", None)
        base.language = getattr(our_args, "language", None) or "en"
        base.audio_duration_s = getattr(our_args, "audio_duration_s", None)
        base.audio_silence_ms = getattr(our_args, "audio_silence_ms", 500)
        base.generated_audio_dir = getattr(our_args, "generated_audio_dir", None)
    else:
        dataset = getattr(our_args, 'dataset', 'random')
        if dataset == 'builtin_mtp_chat' and not our_args.tokenizer:
            raise ValueError('builtin_mtp_chat requires --tokenizer')
        base.dataset_name = dataset
        base.dataset_length_policy = getattr(our_args, 'dataset_length_policy', 'exact')
        base.dataset_input_len_tolerance = getattr(
            our_args, 'dataset_input_len_tolerance', 0.2
        )
        base.dataset_on_bucket_shortage = getattr(
            our_args, 'dataset_on_bucket_shortage', 'error'
        )
        base.dataset_sampling = getattr(our_args, 'dataset_sampling', 'shuffle')

    # ── 流量控制：不限速，仅用 max_concurrency 控制并发 ───────────────────────
    base.request_rate = float('inf')

    # ── 分位数：同时输出 P50/P90/P99 ────────────────────────────────────────
    base.metric_percentiles = '50,90,99'
    base.percentile_metrics = 'ttft,tpot,e2el'

    # ── 强制输出完整长度（ignore_eos）────────────────────────────────────────
    base.ignore_eos = True

    # ── 不自动保存 JSON 结果（由本脚本统一保存 CSV/XLSX）───────────────────────
    base.save_result = False
    base.save_detailed = False
    base.plot_timeline = False
    base.plot_dataset_stats = False

    # ── 第一次运行才做 ready check 和 warmup，后续配置跳过（下面动态设置）────────
    # 初始值：不检查，不预热
    base.ready_check_timeout_sec = 0
    base.num_warmups = 0

    # ── warmup 固定并发/输出参数 ────────────────────────────────────────────
    base.warmup_concurrency = getattr(our_args, 'warmup_concurrency', None)
    base.warmup_output_len = getattr(our_args, 'warmup_output_len', None)

    return base


# ─── 从 serve.py 返回的 result_json 提取 CSV 行 ───────────────────────────────

MAX_SEED_VALUE = 2**32


def validate_seed(seed: int) -> None:
    if not 0 <= seed < MAX_SEED_VALUE:
        raise ValueError(f"--seed 必须满足 0 <= seed < {MAX_SEED_VALUE}，当前值: {seed}")


def derive_config_seed(
    *,
    base_seed: int,
    input_len: int,
    output_len: int,
    parallel_num: int,
    prefix_ratio: float,
    config_index: int,
) -> int:
    key = (
        f"{base_seed}:{input_len}:{output_len}:{parallel_num}:"
        f"{prefix_ratio:.12g}:{config_index}"
    )
    return int(hashlib.sha256(key.encode("utf-8")).hexdigest()[:8], 16)


def effective_config_seed(
    *,
    base_seed: int,
    input_len: int,
    output_len: int,
    parallel_num: int,
    prefix_ratio: float,
    config_index: int,
    vary_seed_by_config: bool,
) -> int:
    validate_seed(base_seed)
    if not vary_seed_by_config:
        return base_seed
    return derive_config_seed(
        base_seed=base_seed,
        input_len=input_len,
        output_len=output_len,
        parallel_num=parallel_num,
        prefix_ratio=prefix_ratio,
        config_index=config_index,
    )


def decide_token_usage_source(*, usage_reported_count: int,
                              tokenizer_fallback_count: int,
                              completed: int,
                              has_tokenizer: bool) -> str:
    """决定每行结果的 token 计数来源（写入 CSV 的 token_source 列）。

    - "usage"：所有成功请求的服务端流式 usage 都上报了 completion_tokens（最可信）
    - "partial_usage"：只有部分成功请求上报 usage，统计混有 fallback
    - "tokenizer_fallback"：服务端未上报，但有本地 tokenizer 可重编码 generated_text 估算
    - "client_estimate"：无 usage/无 tokenizer，只能依赖客户端估算值
    - "none"：没有成功请求
    """
    if completed <= 0:
        return "none"
    if usage_reported_count >= completed:
        return "usage"
    if usage_reported_count > 0:
        return "partial_usage"
    if tokenizer_fallback_count > 0 or has_tokenizer:
        return "tokenizer_fallback"
    return "client_estimate"


def _derive_prefix_suffix_tokens(input_len: int, prefix_ratio: float) -> Tuple[int, int]:
    """Return shared-prefix and unique-suffix lengths within total input_len."""
    if not (0.0 <= prefix_ratio <= 1.0):
        raise ValueError("--prefix-ratio must be between 0.0 and 1.0")
    prefix_tokens = int(input_len * prefix_ratio)
    suffix_tokens = input_len - prefix_tokens
    return prefix_tokens, suffix_tokens


def _is_random_prompt_generation_error(exc: ValueError) -> bool:
    message = str(exc)
    return "unique suffix" in message or "unique prompt collision" in message


def _extract_row(
    result: dict,
    in_len: int,
    out_len: int,
    parallel_num: int,
    epochs: int,
    model: str,
    backend: str,
    prefix_tokens: int = 0,
    prefix_ratio: float = 0.0,
    has_tokenizer: bool = False,
    seed: int = 0,
    *,
    dataset_name: str = "random",
    language: str = "",
) -> dict:
    """从 _serve.main_async() 返回的字典中提取并重命名需要的指标字段。

    重要：avg_input_tokens / avg_output_tokens 取自【存活】的总量字段
    total_input_tokens / total_output_tokens ÷ completed，【不再】回退到 requested 值。
    （serve.py 在 save_detailed=False 时会删除每请求 input_lens/output_lens，
     故不能依赖它们；总量字段不受删除影响。）
    """
    def _f(key: str, default=0.0) -> float:
        v = result.get(key, default)
        return round(float(v), 4) if v is not None else 0.0

    def _i(key: str, default=0) -> int:
        return int(result.get(key, default) or default)

    def _i_any(*keys: str, default=0) -> int:
        for key in keys:
            if result.get(key) is not None:
                return int(result.get(key) or default)
        return default

    def _json_list(key: str) -> str:
        value = result.get(key) or []
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))

    def _rate(numerator: float, denominator_s: float) -> float:
        if denominator_s <= 0:
            return 0.0
        return round(float(numerator) / float(denominator_s), 4)

    completed = _i('completed')
    total_in = _i('total_input_tokens')
    total_out = _i('total_output_tokens')
    total_cached = _i('total_cached_tokens')

    # ── 真实平均 token 数（来自服务端上报的总量）─────────────────────────────
    # 注意：completed==0（全部失败）时回退 0.0，【不】回退 requested 值——
    # 否则会把零成功伪装成"按指定长度输出"，复发 Bug①。
    avg_in = round(total_in / completed, 1) if completed > 0 else 0.0
    avg_out = round(total_out / completed, 1) if completed > 0 else 0.0
    avg_cached_tokens = round(total_cached / completed, 1) if completed > 0 else 0.0
    cache_hit_rate = (
        round(total_cached / total_in * 100, 1) if total_in > 0 else 0.0
    )

    # ── token 计数来源 ────────────────────────────────────────────────────────
    token_source = decide_token_usage_source(
        usage_reported_count=_i('usage_reported_count'),
        tokenizer_fallback_count=_i('tokenizer_fallback_count'),
        completed=completed,
        has_tokenizer=has_tokenizer,
    )

    # ── 长度合规：实测 / 请求 ────────────────────────────────────────────────
    # 基于未取整的真实均值算合规，避免 round(avg_in/out) 在阈值边界引入误差。
    total_input_len = in_len
    raw_avg_in = total_in / completed if completed > 0 else 0.0
    raw_avg_out = total_out / completed if completed > 0 else 0.0
    input_compliance = (
        round(raw_avg_in / total_input_len * 100, 1)
        if total_input_len > 0 else 0.0
    )
    output_compliance = round(raw_avg_out / out_len * 100, 1) if out_len > 0 else 0.0
    finish_reason_length_pct = (
        round(_i('finish_reason_length') / completed * 100, 1)
        if completed > 0 else 0.0
    )

    duration_s = float(result.get('duration', 0.0) or 0.0)
    mean_ttft_ms = float(result.get('mean_ttft_ms', 0.0) or 0.0)
    mean_tpot_ms = float(result.get('mean_tpot_ms', 0.0) or 0.0)
    input_throughput_tok_s = _rate(total_in, duration_s)
    prefill_effective_tok_s = _rate(raw_avg_in, mean_ttft_ms / 1000.0)
    decode_effective_tok_s = _rate(1.0, mean_tpot_ms / 1000.0)
    rtfx = _f('rtfx')
    audio_duration_s_total = round(rtfx * duration_s, 4) if rtfx > 0 else 0.0
    audio_duration_s_avg = (
        round(audio_duration_s_total / completed, 4) if completed > 0 else 0.0
    )

    return {
        # ── 测试配置 ────────────────────────────────
        'model':           model,
        'backend':         backend,
        'dataset_name':    dataset_name,
        'language':        language,
        'input_len':       in_len,           # requested 总输入长度
        'output_len':      out_len,           # requested 输出长度
        'total_input_len': total_input_len,  # 新语义下等于 input_len
        'prefix_ratio':    round(prefix_ratio, 3),
        'prefix_tokens':   prefix_tokens,
        'parallel_num':    parallel_num,
        'epochs':          epochs,
        'num_prompts':     _i('num_prompts', parallel_num * epochs),
        'seed':            seed,
        # ── 请求统计 ────────────────────────────────
        'n_success':           completed,
        'n_failed':            _i('failed'),
        'avg_input_tokens':    avg_in,    # 真实（曾因 Bug①回显 requested，现已修正）
        'avg_output_tokens':   avg_out,   # 真实
        'input_compliance':    input_compliance,
        'output_compliance':   output_compliance,
        'finish_reason_length_pct': finish_reason_length_pct,
        'token_source':        token_source,
        'avg_cached_tokens':   avg_cached_tokens,   # 平均命中缓存的 prompt token 数
        'cache_hit_rate':      cache_hit_rate,       # token 加权缓存命中率 (%) = total_cached/total_in*100
        'cache_hit_rate_metrics': _f('cache_hit_rate_metrics'),
        'cache_hit_tokens_device': _i('cache_hit_tokens_device'),
        'cache_hit_tokens_host': _i('cache_hit_tokens_host'),
        'cache_hit_tokens_storage': _i('cache_hit_tokens_storage'),
        'cache_hit_tokens_storage_mooncake': _i('cache_hit_tokens_storage_mooncake'),
        'avg_gpu_kv_cache_usage': _f('avg_gpu_kv_cache_usage'),
        'peak_gpu_kv_cache_usage': _f('peak_gpu_kv_cache_usage'),
        # ── MTP / Spec Decode ───────────────────────
        'spec_decode_acceptance_rate': _f('spec_decode_acceptance_rate'),
        'spec_decode_system_efficiency': _f('spec_decode_system_efficiency'),
        'spec_decode_num_drafts': _i('spec_decode_num_drafts'),
        'spec_decode_num_accepted_tokens': _i_any(
            'spec_decode_num_accepted_tokens',
            'spec_decode_accepted_tokens',
        ),
        'spec_decode_num_draft_tokens': _i_any(
            'spec_decode_num_draft_tokens',
            'spec_decode_draft_tokens',
        ),
        'spec_decode_per_position_acceptance_rates': _json_list(
            'spec_decode_per_position_acceptance_rates'
        ),
        # ── 吞吐量 ──────────────────────────────────
        'throughput_req_s':   _f('request_throughput'),
        'throughput_tok_s':   _f('output_throughput'),  # 输出 Token 系统吞吐，vLLM 官方口径
        'input_throughput_tok_s': input_throughput_tok_s,
        'prefill_effective_tok_s': prefill_effective_tok_s,
        'decode_effective_tok_s': decode_effective_tok_s,
        # ── TTFT：首 token 时延 (ms) ─────────────────
        'ttft_mean_ms':  _f('mean_ttft_ms'),
        'ttft_p50_ms':   _f('p50_ttft_ms'),
        'ttft_p90_ms':   _f('p90_ttft_ms'),
        'ttft_p99_ms':   _f('p99_ttft_ms'),
        # ── TPOT：每 token 解码时延 (ms) ────────────
        'tpot_mean_ms':  _f('mean_tpot_ms'),
        'tpot_p50_ms':   _f('p50_tpot_ms'),
        'tpot_p90_ms':   _f('p90_tpot_ms'),
        'tpot_p99_ms':   _f('p99_tpot_ms'),
        # ── E2EL：端到端延迟 (ms) ───────────────────
        'e2el_mean_ms':  _f('mean_e2el_ms'),
        'e2el_p50_ms':   _f('p50_e2el_ms'),
        'e2el_p90_ms':   _f('p90_e2el_ms'),
        'e2el_p99_ms':   _f('p99_e2el_ms'),
        # ── 其他 ────────────────────────────────────
        'audio_duration_s_total': audio_duration_s_total,
        'audio_duration_s_avg':   audio_duration_s_avg,
        'rtfx':          rtfx,
        'duration_s':    _f('duration'),
    }


# ─── 结果保存 ──────────────────────────────────────────────────────────────────

CSV_HEADERS = [
    # ── 基本信息 ──────────────────────────────────
    'model', 'backend', 'dataset_name', 'language',
    'input_len', 'output_len', 'total_input_len', 'prefix_ratio', 'prefix_tokens',
    'parallel_num', 'epochs', 'num_prompts', 'seed',
    'n_success', 'n_failed',
    'avg_input_tokens', 'avg_output_tokens',
    # ── 吞吐量（前移）──────────────────────────────
    'throughput_req_s', 'throughput_tok_s', 'input_throughput_tok_s',
    'prefill_effective_tok_s', 'decode_effective_tok_s',
    # ── TTFT（前移）────────────────────────────────
    'ttft_mean_ms', 'ttft_p50_ms', 'ttft_p90_ms', 'ttft_p99_ms',
    # ── 其他 ──────────────────────────────────────
    'input_compliance', 'output_compliance',
    'finish_reason_length_pct', 'token_source',
    'avg_cached_tokens', 'cache_hit_rate',
    'cache_hit_rate_metrics',
    'cache_hit_tokens_device', 'cache_hit_tokens_host',
    'cache_hit_tokens_storage', 'cache_hit_tokens_storage_mooncake',
    'avg_gpu_kv_cache_usage', 'peak_gpu_kv_cache_usage',
    'spec_decode_acceptance_rate', 'spec_decode_system_efficiency',
    'spec_decode_num_drafts', 'spec_decode_num_accepted_tokens',
    'spec_decode_num_draft_tokens', 'spec_decode_per_position_acceptance_rates',
    'tpot_mean_ms', 'tpot_p50_ms', 'tpot_p90_ms', 'tpot_p99_ms',
    'e2el_mean_ms', 'e2el_p50_ms', 'e2el_p90_ms', 'e2el_p99_ms',
    'audio_duration_s_total', 'audio_duration_s_avg', 'rtfx',
    'duration_s',
]

CSV_HEADERS_ZH = [
    # ── 基本信息 ──
    '模型', '接口类型', '数据集', '语言',
    '输入长度(token)', '输出长度(token)', '总输入长度(token)', '前缀比例', '前缀tokens数',
    '并发数', '测试轮数', '总请求数', '随机种子',
    '成功请求数', '失败请求数',
    '平均实际输入tokens', '平均实际输出tokens',
    # ── 吞吐量（前移）──
    '请求吞吐(req/s)', '输出Token系统吞吐(tok/s)', '输入Token系统吞吐(tok/s)',
    'Prefill有效速率(tok/s)', 'Decode有效速率(tok/s)',
    # ── TTFT（前移）──
    'TTFT均值(ms)', 'TTFT_P50(ms)', 'TTFT_P90(ms)', 'TTFT_P99(ms)',
    # ── 其他 ──
    '输入长度合规(%)', '输出长度合规(%)',
    'length停止占比(%)', 'token来源',
    '平均缓存命中tokens', '缓存命中率(%)',
    'Metrics缓存命中率(%)',
    'Device缓存命中tokens', 'Host缓存命中tokens',
    'Storage缓存命中tokens', 'Mooncake缓存命中tokens',
    '平均GPU KV缓存占用率(%)', '峰值GPU KV缓存占用率(%)',
    'SpecDecode接受率(%)', 'SpecDecode系统效率',
    'SpecDecode草稿轮数', 'SpecDecode接受tokens数',
    'SpecDecode草稿tokens数', 'SpecDecode分位置接受率(%)',
    'TPOT均值(ms)', 'TPOT_P50(ms)', 'TPOT_P90(ms)', 'TPOT_P99(ms)',
    'E2EL均值(ms)', 'E2EL_P50(ms)', 'E2EL_P90(ms)', 'E2EL_P99(ms)',
    '音频总时长(s)', '平均音频时长(s)', 'RTFx',
    '测试耗时(s)',
]


def save_csv(rows: List[dict], path: str) -> None:
    """将结果列表保存为 CSV 文件（UTF-8-BOM 编码，Excel 可直接打开）"""
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)
    logger.info("✓ CSV 结果已保存: %s", path)


def save_xlsx(rows: List[dict], path: str) -> None:
    """将结果列表保存为带格式的 XLSX 文件"""
    if not _HAS_OPENPYXL:
        logger.error("保存 XLSX 需要 openpyxl，请执行: pip install openpyxl")
        return

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '基准测试结果'
    ws.freeze_panes = 'A3'  # 冻结前两行（表头）

    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill(fill_type='solid', fgColor='1F497D')
    sub_font  = Font(bold=False, color='FFFFFF', size=9)
    sub_fill  = PatternFill(fill_type='solid', fgColor='366092')
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, h in enumerate(CSV_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
    for col, h in enumerate(CSV_HEADERS_ZH, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = sub_font
        c.fill = sub_fill
        c.alignment = center

    fill_a = PatternFill(fill_type='solid', fgColor='F2F7FF')
    fill_b = PatternFill(fill_type='solid', fgColor='FFFFFF')
    color_toggle: Dict[Tuple, int] = {}
    color_idx = 0

    for row_idx, row in enumerate(rows, 3):
        key = (row.get('input_len'), row.get('output_len'))
        if key not in color_toggle:
            color_toggle[key] = color_idx % 2
            color_idx += 1
        row_fill = fill_a if color_toggle[key] == 0 else fill_b
        for col, key_name in enumerate(CSV_HEADERS, 1):
            val = row.get(key_name, '')
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill = row_fill
            c.alignment = Alignment(horizontal='right' if isinstance(val, float) else 'center')

    n_rows = len(rows) + 2
    for col in range(1, len(CSV_HEADERS) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or '')) for r in range(1, n_rows + 1)),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 30)

    # 第二页：指标说明
    ws2 = wb.create_sheet(title='指标说明')
    explain = [
        ('指标', '含义', '计算方式（来自 vllm_bench/serve.py）'),
        ('TTFT', '首 Token 时延 (Time-To-First-Token)', '首个 token 到达时间 − 请求发出时间'),
        ('TPOT', '每输出 Token 时延 (Time-Per-Output-Token)', '(E2EL − TTFT) ÷ (output_tokens − 1)'),
        ('E2EL', '端到端延迟 (End-to-End Latency)', '最后一个 token 到达时间 − 请求发出时间'),
        ('throughput_req_s', '请求吞吐量', 'completed ÷ benchmark_duration'),
        ('throughput_tok_s', '输出 Token 系统吞吐量', 'total_output_tokens ÷ benchmark_duration（vLLM 官方 output_throughput 口径）'),
        ('input_throughput_tok_s', '输入 Token 系统吞吐量', 'total_input_tokens ÷ benchmark_duration'),
        ('prefill_effective_tok_s', 'Prefill 有效速率', 'avg_input_tokens ÷ mean_TTFT_s；TTFT 含排队、调度和首 token'),
        ('decode_effective_tok_s', 'Decode 有效速率', '1 ÷ mean_TPOT_s；基于 TPOT 的 next-token decode 近似速率'),
        ('avg_cached_tokens', '平均缓存命中 tokens', 'total_cached_tokens ÷ completed（服务端 usage.cached_tokens 累计）'),
        ('cache_hit_rate', '缓存命中率(%)', 'total_cached_tokens ÷ total_input_tokens × 100（token 加权；仅服务端开启 prefix caching 时非零）'),
        ('cache_hit_rate_metrics', 'Metrics 缓存命中率(%)', 'SGLang /metrics 中 cached_tokens_total 差分 ÷ prompt_tokens_total 差分 × 100'),
        ('cache_hit_tokens_device', 'Device 缓存命中 tokens', 'SGLang /metrics 中 cache_source=device 的 cached_tokens_total 差分'),
        ('cache_hit_tokens_host', 'Host 缓存命中 tokens', 'SGLang /metrics 中 cache_source=host 的 cached_tokens_total 差分'),
        ('cache_hit_tokens_storage', 'Storage 缓存命中 tokens', '所有 cache_source=storage_* 的 cached_tokens_total 差分之和'),
        ('cache_hit_tokens_storage_mooncake', 'Mooncake 缓存命中 tokens', 'SGLang /metrics 中 cache_source=storage_mooncake 的 cached_tokens_total 差分'),
        ('avg_gpu_kv_cache_usage', '平均 GPU KV cache 占用率(%)', 'benchmark 运行期间定期采样 /metrics 中 GPU KV cache usage 后求平均值'),
        ('peak_gpu_kv_cache_usage', '峰值 GPU KV cache 占用率(%)', 'benchmark 运行期间定期采样 /metrics 中 GPU KV cache usage 后取峰值'),
        ('spec_decode_acceptance_rate', 'Spec Decode 接受率(%)', 'accepted_tokens ÷ draft_tokens × 100（来自 /metrics 差分）'),
        ('spec_decode_system_efficiency', 'Spec Decode 系统效率', '接受 tokens 相对 draft tokens 的服务端效率指标'),
        ('spec_decode_num_drafts', 'Spec Decode 草稿轮数', 'num_drafts_total 差分'),
        ('spec_decode_num_accepted_tokens', 'Spec Decode 接受 tokens 数', 'num_accepted_tokens_total 差分'),
        ('spec_decode_num_draft_tokens', 'Spec Decode 草稿 tokens 数', 'num_draft_tokens_total 差分'),
        ('spec_decode_per_position_acceptance_rates', '分位置接受率(%)', '各 speculative position 的接受率列表'),
        ('P50/P90/P99', '百分位数', 'P90 表示 90% 请求低于该延迟值'),
        ('', '', ''),
        ('并发控制说明', '',
         'parallel_num → max_concurrency；request_rate=inf（不限速，仅限并发数）'),
        ('请求总数说明', '',
         'num_prompts = parallel_num × epochs（每轮 parallel_num 个并发请求）'),
    ]
    tf = Font(bold=True, color='FFFFFF', size=10)
    tfl = PatternFill(fill_type='solid', fgColor='1F497D')
    for ri, rd in enumerate(explain, 1):
        for ci, val in enumerate(rd, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            if ri == 1:
                c.font = tf
                c.fill = tfl
            c.alignment = Alignment(wrap_text=True, vertical='center')
    for ci, w in zip(range(1, 4), [22, 45, 55]):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    wb.save(path)
    logger.info("✓ XLSX 结果已保存: %s", path)


# ─── 主循环 ───────────────────────────────────────────────────────────────────

def _run_all(our_args: argparse.Namespace) -> List[dict]:
    """
    遍历所有 (input_len, output_len, parallel_num) 组合，
    为每组构建 serve.py 所需 args，调用 _serve.main_async()，
    收集并返回所有指标行。
    """
    backend = getattr(our_args, "backend", "openai")
    is_asr_backend = backend == ASR_BACKEND
    if is_asr_backend and not getattr(our_args, "dataset_path", None):
        raise ValueError("--backend openai-audio requires --dataset-path")

    validate_seed(our_args.seed)
    vary_seed_by_config = not our_args.no_vary_seed_by_config

    # 构建基础 Namespace（含全部 serve.py 默认值）
    base = _build_base_args(our_args)
    if is_asr_backend:
        if list(our_args.input_lens) != [0] or our_args.prefix_ratio != 0.0:
            logger.warning(
                "ASR backend openai-audio ignores --input-lens and "
                "--prefix-ratio; using input_lens=[0], prefix_ratio=0.0"
            )
    else:
        _derive_prefix_suffix_tokens(1, our_args.prefix_ratio)
    model = our_args.served_model_name or our_args.model

    # 构建 (in_len, out_len) 测试对
    in_lens: List[int] = [0] if is_asr_backend else our_args.input_lens
    out_lens: List[int] = our_args.output_lens

    if our_args.cross_product:
        io_pairs = list(itertools.product(in_lens, out_lens))
        logger.info("笛卡尔积模式: %d × %d = %d 组", len(in_lens), len(out_lens), len(io_pairs))
    elif is_asr_backend:
        io_pairs = [(0, out_len) for out_len in out_lens]
    else:
        if len(out_lens) == 1:
            out_lens = out_lens * len(in_lens)
        elif len(out_lens) != len(in_lens):
            logger.error(
                "--input-lens 有 %d 个值，--output-lens 有 %d 个值。"
                "需要数量相同、或 --output-lens 只给 1 个（广播）、或加 --cross-product",
                len(in_lens), len(out_lens),
            )
            sys.exit(1)
        io_pairs = list(zip(in_lens, out_lens))

    sorted_parallels = sorted(our_args.parallel_nums)

    logger.info("=" * 65)
    logger.info("测试参数汇总:")
    endpoint_desc = our_args.base_url or f"{our_args.host}:{our_args.port}"
    logger.info("  服务      : %s  backend=%s", endpoint_desc, our_args.backend)
    logger.info("  模型      : %s", model)
    logger.info("  (输入,输出) 组合: %s", io_pairs)
    logger.info("  并发数    : %s", sorted_parallels)
    logger.info("  每组轮数  : %d  → num_prompts = parallel × epochs", our_args.epochs)
    logger.info("  随机种子  : base=%d  vary_by_config=%s",
                our_args.seed, vary_seed_by_config)
    logger.info("  并发模型  : 滑动窗口（Semaphore=parallel_num, request_rate=inf）")
    logger.info("             每组一次性提交 parallel×epochs 个任务，Semaphore 保证")
    logger.info("             最多 parallel_num 个请求同时在途，非严格批次轮转")
    logger.info("  配置间隔  : %.1fs", our_args.sleep_between)
    logger.info("  底层引擎  : vllm_bench/serve.py（shim via run_bench_serve.py）")
    logger.info("=" * 65)

    all_rows: List[dict] = []
    config_count = 0
    total_configs = len(io_pairs) * len(sorted_parallels)
    is_first_run = True  # 仅第一次做 ready check

    for in_len, out_len in io_pairs:
        skip_higher_parallel = False

        for parallel_num in sorted_parallels:
            config_count += 1

            if skip_higher_parallel:
                logger.warning(
                    "[%d/%d] 跳过 input=%d output=%d parallel=%d（已超出约束）",
                    config_count, total_configs, in_len, out_len, parallel_num,
                )
                continue

            prefix_ratio = 0.0 if is_asr_backend else our_args.prefix_ratio
            cfg_input_len = 0 if is_asr_backend else in_len
            if is_asr_backend:
                prefix_tokens = 0
                suffix_tokens = 0
            else:
                prefix_tokens, suffix_tokens = _derive_prefix_suffix_tokens(
                    in_len, prefix_ratio
                )
            effective_seed = effective_config_seed(
                base_seed=our_args.seed,
                input_len=cfg_input_len,
                output_len=out_len,
                parallel_num=parallel_num,
                prefix_ratio=prefix_ratio,
                config_index=config_count,
                vary_seed_by_config=vary_seed_by_config,
            )
            prefix_detail = ""
            if not is_asr_backend:
                prefix_detail = (
                    f"  total_input={in_len} prefix={prefix_tokens}tok"
                    f"({prefix_ratio * 100:.0f}%) suffix={suffix_tokens}tok"
                )

            logger.info(
                "\n%s\n[%d/%d] 开始测试: input=%d, output=%d, parallel=%d, "
                "num_prompts=%d (=%d×%d epochs), seed=%d%s\n%s",
                "─" * 65,
                config_count, total_configs,
                cfg_input_len, out_len, parallel_num,
                parallel_num * our_args.epochs, parallel_num, our_args.epochs,
                effective_seed,
                prefix_detail,
                "─" * 65,
            )

            # 复制 base args，按本次配置覆盖可变字段
            cfg = copy.copy(base)
            cfg.input_len       = cfg_input_len  # serve.py 内部映射到数据集输入长度
            cfg.output_len      = out_len       # serve.py 内部映射到数据集输出长度
            if is_asr_backend:
                cfg.custom_output_len = out_len
            cfg.max_concurrency = parallel_num  # 最大并发数
            cfg.num_prompts     = parallel_num * our_args.epochs  # 总请求数
            cfg.seed            = effective_seed

            # input_len 表示总 prompt token 预算；prefix_tokens 是其中共享前缀部分。
            # run_bench_serve.py 会生成 shared_prefix + unique_suffix，总长约等于 input_len。
            cfg.random_prefix_len = prefix_tokens

            # 第一次运行：做 ready check（超时 600s）；后续跳过（设为 0）
            if is_first_run:
                cfg.ready_check_timeout_sec = 600
                # 固定并发预热：warmup_requests 表示轮数；
                # warmup_concurrency 设了则每轮发满该并发。
                warmup_cc = getattr(our_args, 'warmup_concurrency', None)
                cfg.num_warmups = our_args.warmup_requests * (
                    warmup_cc if warmup_cc is not None else 1
                )
                is_first_run = False
            else:
                cfg.ready_check_timeout_sec = 0
                cfg.num_warmups = 0

            # 调用 vllm_bench/serve.py 的核心测试逻辑
            try:
                result = asyncio.run(_serve.main_async(cfg))
            except ValueError as exc:
                if _is_random_prompt_generation_error(exc):
                    raise
                logger.error("测试失败 (input=%d, output=%d, parallel=%d): %s",
                             in_len, out_len, parallel_num, exc)
                continue
            except Exception as exc:
                logger.error("测试失败 (input=%d, output=%d, parallel=%d): %s",
                             in_len, out_len, parallel_num, exc)
                continue

            # 提取汇总指标
            row = _extract_row(result, cfg_input_len, out_len, parallel_num,
                               our_args.epochs, model, our_args.backend,
                               prefix_tokens=prefix_tokens,
                               prefix_ratio=prefix_ratio,
                               has_tokenizer=(
                                   False if is_asr_backend else bool(our_args.tokenizer)
                               ),
                               seed=effective_seed,
                               dataset_name=getattr(cfg, "dataset_name", "random"),
                               language=(
                                   getattr(cfg, "language", "") if is_asr_backend else ""
                               ))
            all_rows.append(row)

            # 打印摘要行
            logger.info(
                "  ✓ 结果: out_sys_tok/s=%.1f  in_sys_tok/s=%.1f  "
                "prefill_eff_tok/s=%.1f  decode_eff_tok/s=%.1f  "
                "TTFT_mean=%.1fms  TTFT_p90=%.1fms  TPOT_mean=%.3fms  "
                "E2EL_mean=%.1fms  成功=%d",
                row['throughput_tok_s'],
                row['input_throughput_tok_s'],
                row['prefill_effective_tok_s'],
                row['decode_effective_tok_s'],
                row['ttft_mean_ms'], row['ttft_p90_ms'],
                row['tpot_mean_ms'],
                row['e2el_mean_ms'],
                row['n_success'],
            )

            # 实时保存（防止中途中断丢失数据）
            if our_args.output_csv:
                save_csv(all_rows, our_args.output_csv)

            # 约束检查：TTFT 均值超限则跳过同组更高并发
            if our_args.max_ttft_ms is not None and row['ttft_mean_ms'] > our_args.max_ttft_ms:
                logger.warning(
                    "  TTFT_mean (%.1fms) > 约束上限 (%.1fms)，"
                    "跳过 input=%d output=%d 的更高并发测试",
                    row['ttft_mean_ms'], our_args.max_ttft_ms, in_len, out_len,
                )
                skip_higher_parallel = True

            # 约束检查：输出吞吐低于下限（通常 parallel=1 时触发）则跳过更高并发
            if (not skip_higher_parallel
                    and our_args.min_throughput_tok_s is not None
                    and row['throughput_tok_s'] < our_args.min_throughput_tok_s):
                logger.warning(
                    "  throughput_tok_s (%.1f) < 约束下限 (%.1f tok/s)，"
                    "跳过 input=%d output=%d 的更高并发测试",
                    row['throughput_tok_s'], our_args.min_throughput_tok_s, in_len, out_len,
                )
                skip_higher_parallel = True

            # 约束检查：输出长度合规偏低（服务端未按指定长度输出）则跳过更高并发
            compliance_frac = row['output_compliance'] / 100.0
            if (not skip_higher_parallel
                    and out_len > 0
                    and row['n_success'] > 0
                    and our_args.min_output_compliance is not None
                    and compliance_frac < our_args.min_output_compliance):
                logger.warning(
                    "  output_compliance (%.1f%%) < 约束下限 (%.1f%%)，"
                    "token_source=%s —— 服务端可能未按指定长度输出"
                    "（ignore_eos 未生效/被截断），"
                    "跳过 input=%d output=%d 的更高并发测试",
                    row['output_compliance'], our_args.min_output_compliance * 100,
                    row['token_source'], in_len, out_len,
                )
                skip_higher_parallel = True

            # 配置间隔等待
            if our_args.sleep_between > 0 and config_count < total_configs:
                logger.info("  等待 %.1fs 让服务稳定...", our_args.sleep_between)
                time.sleep(our_args.sleep_between)

    return all_rows


# ─── CLI 参数解析 ──────────────────────────────────────────────────────────────

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            '多配置 vLLM OpenAI 接口批量基准测试\n'
            '（复用 run_bench_serve.py shim + vllm_bench/serve.py 测试核心）'
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 多组输入输出长度 + 多并发
  python run_bench_multi.py \\
      --host 127.0.0.1 --port 8000 \\
      --model Qwen2.5-7B-Instruct \\
      --input-lens 128 512 1024 2048 \\
      --output-lens 128 256 512 512 \\
      --parallel-nums 1 4 8 16 \\
      --epochs 3 \\
      --output-csv results/bench.csv \\
      --output-xlsx results/bench.xlsx

  # 笛卡尔积模式（所有 input × output 组合）
  python run_bench_multi.py \\
      --host 127.0.0.1 --port 8000 --model mymodel \\
      --input-lens 512 1024 --output-lens 128 512 --cross-product \\
      --parallel-nums 1 8 16 --epochs 5
        """,
    )

    # ── 服务连接 ────────────────────────────────────────────────────────────
    conn = p.add_argument_group('服务连接')
    conn.add_argument('--base-url', default=None,
                      help='完整服务 URL（与 --host/--port 二选一）。'
                           '末尾不加斜杠，需已包含版本路径（如 /v1）。'
                           '例: https://aicp.example.com/openai/api/v1/openai/v1 。'
                           '最终 api_url = base-url + /completions（或 /chat/completions）')
    conn.add_argument('--host', default='127.0.0.1',
                      help='vLLM 服务 IP 地址（--base-url 未设置时使用，默认: 127.0.0.1）')
    conn.add_argument('--port', type=int, default=8000,
                      help='vLLM 服务端口号（--base-url 未设置时使用，默认: 8000）')
    conn.add_argument('--insecure', action='store_true', default=False,
                      help='跳过 HTTPS 证书验证（仅在使用 --base-url 且服务证书不受信时使用）')
    conn.add_argument('--model', required=True,
                      help='模型名称，需与服务端 --model 完全一致')
    conn.add_argument('--served-model-name', default=None,
                      help='服务端挂载的模型别名（若与 --model 不同时指定）')
    conn.add_argument('--backend', choices=['openai', 'openai-chat', ASR_BACKEND],
                      default='openai',
                      help='请求接口类型: openai=/v1/completions, '
                           'openai-chat=/v1/chat/completions, '
                           'openai-audio=/v1/audio/transcriptions（默认: openai）')
    conn.add_argument('--api-key', default=None,
                      help='API 认证密钥（若服务端开启鉴权时指定）')

    # ── 测试配置 ────────────────────────────────────────────────────────────
    bench = p.add_argument_group('测试配置')
    bench.add_argument('--input-lens', type=int, nargs='+', default=[512],
                       help='输入 token 长度列表，支持多个值（空格分隔）。'
                            'ASR/openai-audio 会忽略该参数并使用 0')
    bench.add_argument('--output-lens', type=int, nargs='+', default=[128],
                       help='输出 token 长度列表；数量需与 --input-lens 一致，'
                            '或只给 1 个值（广播到所有输入长度）')
    bench.add_argument('--cross-product', action='store_true', default=False,
                       help='笛卡尔积模式：测试所有 input-lens × output-lens 的组合')
    bench.add_argument('--parallel-nums', type=int, nargs='+', default=[1, 4, 8],
                       help='并发请求数列表，每个值独立测试（空格分隔）'
                            '；内部映射到 serve.py 的 --max-concurrency')
    bench.add_argument('--epochs', type=int, default=3,
                       help='每组配置的重复轮数；总请求数 = parallel × epochs（默认: 3）')
    bench.add_argument('--sleep-between', type=float, default=2.0,
                       help='每组配置测试后的等待时间（秒，让服务 KV cache 释放，默认: 2.0）')
    bench.add_argument('--warmup-requests', type=int, default=1,
                       help='首次测试前的预热轮数；设 warmup-concurrency 后总预热请求数'
                            '= warmup_requests × warmup_concurrency（默认: 1）')
    bench.add_argument('--warmup-concurrency', type=int, default=None,
                       help='warmup 每轮固定并发数（仅首次预热生效；默认 None=跟随该档并发）')
    bench.add_argument('--warmup-output-len', type=int, default=None,
                       help='warmup 请求输出长度（默认 None=跟随该档输出；设短值省 decode 时间）')
    bench.add_argument('--dataset-name', default=None,
                       help='ASR 使用的数据集类型；openai-audio 默认 custom_audio')
    bench.add_argument('--dataset-path', default=None,
                       help='ASR JSONL 数据集路径；custom_audio 需要')
    bench.add_argument('--language', default='en',
                       help='OpenAI Audio transcription language 参数，默认 en')
    bench.add_argument('--audio-duration-s', type=float, default=None,
                       help='ASR/custom_audio: 运行时拼接源音频，生成每条请求的目标音频时长（秒）。'
                            '不设置时直接使用数据集原始音频')
    bench.add_argument('--audio-silence-ms', type=int, default=500,
                       help='ASR/custom_audio: 动态拼接片段之间插入的静音长度（毫秒，默认 500）')
    bench.add_argument('--generated-audio-dir', default=None,
                       help='ASR/custom_audio: 动态拼接音频与 manifest 输出目录；'
                            '不设置时使用临时目录')
    bench.add_argument('--prefix-ratio', type=float, default=0.0,
                       help='前缀缓存比例 [0.0~1.0]（默认: 0.0 = 不使用共享前缀）。'
                            '取值 0.5 表示每个请求中有 50%% 的 input_len 作为共享前缀。'
                            '所有请求共享同一段前缀文本（固定生成一次），'
                            '后缀包含 request-index token 加随机 tail，用于保证非 full-prefix 请求差异。'
                            '实际 prompt_len ≈ input_len；共享前缀 token 数约为 input_len × prefix_ratio。'
                            '用于对比 prefix caching 开启/关闭对延迟/吞吐的影响。'
                            'ASR/openai-audio 会忽略该参数并使用 0.0。')
    bench.add_argument('--dataset', default='random',
                       choices=['random', 'builtin_mtp_chat'],
                       help='请求数据集（默认 random；builtin_mtp_chat 使用内置真实风格 MTP chat prompt）')
    bench.add_argument('--dataset-length-policy', default='exact',
                       choices=['exact', 'bucket'],
                       help='dataset 输入长度策略：exact 精确匹配，bucket 按容忍度筛选')
    bench.add_argument('--dataset-input-len-tolerance', type=float, default=0.2,
                       help='dataset-length-policy=bucket 时的 input_len 容忍度')
    bench.add_argument('--dataset-on-bucket-shortage', default='error',
                       choices=['error'],
                       help='bucket 无候选样本时的处理策略')
    bench.add_argument('--dataset-sampling', default='shuffle',
                       choices=['shuffle', 'round_robin'],
                       help='内置数据集采样顺序')
    bench.add_argument('--seed', type=int, default=0,
                       help='随机种子基值。默认每个配置会基于该值派生独立 seed（默认: 0）')
    bench.add_argument('--no-vary-seed-by-config', action='store_true', default=False,
                       help='兼容旧行为：所有配置复用 --seed，不再按配置派生独立 seed')

    # ── Tokenizer ────────────────────────────────────────────────────────────
    tok = p.add_argument_group('Tokenizer（可选）')
    tok.add_argument('--tokenizer', default=None,
                     help='HuggingFace tokenizer 路径，用于精确生成指定长度的随机 token；'
                          '不指定则由 serve.py 以近似字符串模式生成（skip_tokenizer_init=True）')

    # ── 约束过滤 ────────────────────────────────────────────────────────────
    limit = p.add_argument_group('约束过滤（可选）')
    limit.add_argument('--max-ttft-ms', type=float, default=None,
                       help='最大允许的 TTFT 均值（ms）；超过后跳过该 (input,output) '
                            '组合的更高并发测试')
    limit.add_argument('--min-throughput-tok-s', type=float, default=None,
                       help='最低输出 Token 系统吞吐量阈值（tok/s）；低于此值时跳过该 '
                            '(input,output) 组合的更高并发测试。\n'
                            '通常在 parallel=1（单用户）时触发：若单用户吞吐已低于预期，'
                            '说明服务在该请求规格下性能不达标，无需继续测试更高并发。')
    limit.add_argument('--min-output-compliance', type=float, default=0.95,
                       help='最低输出长度合规比例（0~1，avg_output_measured/requested）。'
                            '低于此值告警并跳过该 (input,output) 组合的更高并发测试。'
                            'ignore_eos 生效时合规应≈1.0；偏低通常意味服务端提前停止'
                            '（如 ignore_eos 未生效/被截断）。默认 0.95。')

    # ── 输出配置 ────────────────────────────────────────────────────────────
    out = p.add_argument_group('输出配置')
    out.add_argument('--output-csv', default='results/bench_multi.csv',
                     help='CSV 结果文件路径（UTF-8-BOM，Excel 可直接打开；默认: results/bench_multi.csv）')
    out.add_argument('--output-xlsx', default=None,
                     help='XLSX 结果文件路径（需安装 openpyxl；不指定则不保存 XLSX）')
    out.add_argument('--result-dir', default=None,
                     help='统一指定输出目录，CSV/XLSX 文件名自动放入此目录')

    return p


def _parse_args() -> argparse.Namespace:
    p = build_arg_parser()
    return p.parse_args()


# ─── 入口 ─────────────────────────────────────────────────────────────────────

def main() -> None:
    args = _parse_args()

    # 处理 result-dir
    if args.result_dir:
        os.makedirs(args.result_dir, exist_ok=True)
        if args.output_csv and not os.path.isabs(args.output_csv):
            args.output_csv = os.path.join(args.result_dir, os.path.basename(args.output_csv))
        if args.output_xlsx and not os.path.isabs(args.output_xlsx):
            args.output_xlsx = os.path.join(args.result_dir, os.path.basename(args.output_xlsx))

    rows = _run_all(args)

    if not rows:
        logger.error("未收集到任何有效测试结果，请检查服务连接和参数配置")
        sys.exit(1)

    # 最终保存
    if args.output_csv:
        save_csv(rows, args.output_csv)
    if args.output_xlsx:
        save_xlsx(rows, args.output_xlsx)

    # 终端汇总表
    print()
    print('=' * 120)
    print(
        f"{'输入':>6} {'输出':>6} {'并发':>5} "
        f"{'out_sys':>10} {'in_sys':>10} {'prefill':>10} {'decode':>10} {'req/s':>8} "
        f"{'命中%':>8} "
        f"{'TTFT均值':>10} {'TTFT_P90':>10} "
        f"{'TPOT均值':>10} {'E2EL均值':>10} {'成功':>6}"
    )
    print('-' * 120)
    for r in rows:
        print(
            f"{r['input_len']:>6} {r['output_len']:>6} {r['parallel_num']:>5} "
            f"{r['throughput_tok_s']:>10.1f} "
            f"{r['input_throughput_tok_s']:>10.1f} "
            f"{r['prefill_effective_tok_s']:>10.1f} "
            f"{r['decode_effective_tok_s']:>10.1f} "
            f"{r['throughput_req_s']:>8.3f} "
            f"{r['cache_hit_rate']:>8.1f} "
            f"{r['ttft_mean_ms']:>10.1f} {r['ttft_p90_ms']:>10.1f} "
            f"{r['tpot_mean_ms']:>10.3f} {r['e2el_mean_ms']:>10.1f} {r['n_success']:>6}"
        )
    print('=' * 120)
    if args.output_csv:
        print(f"CSV  → {args.output_csv}")
    if args.output_xlsx:
        print(f"XLSX → {args.output_xlsx}")
    print()


if __name__ == '__main__':
    main()
