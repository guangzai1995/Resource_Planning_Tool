"""端到端冒烟：monkeypatch serve.main_async 返回固定结果，走 run_bench_multi
的提取与落盘路径，断言 CSV 的 avg/合规列真实。"""
import argparse
import csv
import logging

import pytest

import run_bench_multi as m

# 注意：run_bench_multi 内部用 importlib 把 run_bench_serve.py 加载为独立模块
# （'_run_bench_serve_shims'），所以必须 patch m._serve（即 run_bench_multi 实际
# 调用的那个 serve 模块对象），而不是 import run_bench_serve 得到的另一个实例。
serve = m._serve


def _fake_result(in_len, out_len, completed, *, undergen=False):
    """构造 serve.main_async 风格的 result dict。undergen=True 表示服务端少生成。"""
    real_out = (out_len // 2) if undergen else out_len
    return {
        "duration": 2.0,
        "completed": completed, "failed": 0,
        "total_input_tokens": in_len * completed,
        "total_output_tokens": real_out * completed,
        "request_throughput": completed / 2.0,
        "output_throughput": real_out * completed / 2.0,
        "total_token_throughput": (in_len + real_out) * completed / 2.0,
        "usage_reported_count": completed,
        "finish_reason_length": completed if not undergen else 0,
        "num_prompts": completed,
        "mean_ttft_ms": 50.0, "p50_ttft_ms": 50.0, "p90_ttft_ms": 60.0, "p99_ttft_ms": 70.0,
        "mean_tpot_ms": 30.0, "p50_tpot_ms": 30.0, "p90_tpot_ms": 31.0, "p99_tpot_ms": 32.0,
        "mean_e2el_ms": 1000.0, "p50_e2el_ms": 1000.0, "p90_e2el_ms": 1100.0, "p99_e2el_ms": 1200.0,
    }


def _run_to_rows(monkeypatch, results_seq):
    """把 _serve.main_async 替换为依次返回 results_seq 的假实现，跑 _run_all。"""
    it = iter(results_seq)

    async def _fake_main_async(cfg):
        return next(it)

    monkeypatch.setattr(serve, "main_async", _fake_main_async)

    our_args = argparse.Namespace(
        model="m", served_model_name=None, backend="openai-chat",
        base_url=None, host="127.0.0.1", port=8000, insecure=False, api_key=None,
        tokenizer="/some/tok",  # has_tokenizer=True
        input_lens=[128], output_lens=[8], cross_product=False,
        parallel_nums=[1], epochs=1, sleep_between=0, warmup_requests=0,
        prefix_ratio=0.0,
        seed=0, no_vary_seed_by_config=False,
        max_ttft_ms=None, min_throughput_tok_s=None, min_output_compliance=0.95,
        output_csv=None, output_xlsx=None, result_dir=None,
    )
    return m._run_all(our_args)


def test_csv_records_real_avg_and_compliance(tmp_path, monkeypatch):
    rows = _run_to_rows(monkeypatch, [_fake_result(128, 8, 3)])
    assert rows[0]["avg_output_tokens"] == 8.0          # 真实，非 requested 回显
    assert rows[0]["output_compliance"] == 100.0
    assert rows[0]["token_source"] == "usage"

    csv_path = str(tmp_path / "bench.csv")
    m.save_csv(rows, csv_path)
    with open(csv_path, encoding="utf-8-sig") as f:
        data = list(csv.DictReader(f))
    assert len(data) == 1
    assert float(data[0]["avg_output_tokens"]) == 8.0
    assert float(data[0]["output_compliance"]) == 100.0
    assert data[0]["token_source"] == "usage"
    assert int(data[0]["total_input_len"]) == 128


def test_csv_flags_undergeneration(tmp_path, monkeypatch):
    rows = _run_to_rows(monkeypatch, [_fake_result(128, 8, 3, undergen=True)])
    assert rows[0]["avg_output_tokens"] == 4.0          # 服务端只生成一半
    assert rows[0]["output_compliance"] == 50.0         # 4/8
    assert rows[0]["finish_reason_length_pct"] == 0.0   # 都不是 length 停止


def test_run_all_prefix_ratio_uses_total_input_budget(monkeypatch, tmp_path, caplog):
    seen_cfgs = []

    async def fake_main_async(cfg):
        seen_cfgs.append(cfg)
        return _fake_result(128, 8, 1)

    monkeypatch.setattr(serve, "main_async", fake_main_async)
    caplog.set_level(logging.INFO, logger=m.logger.name)

    args = argparse.Namespace(
        model="m", served_model_name=None, backend="openai",
        base_url="http://x/v1", host="127.0.0.1", port=8000,
        insecure=False, api_key=None, tokenizer="/some/tok",
        input_lens=[128], output_lens=[8], cross_product=False,
        parallel_nums=[1], epochs=1, sleep_between=0,
        warmup_requests=0, prefix_ratio=0.8,
        seed=0, no_vary_seed_by_config=False,
        output_csv=str(tmp_path / "out.csv"), output_xlsx=None,
        result_dir=None, max_ttft_ms=None, min_throughput_tok_s=None,
        min_output_compliance=0.0,
    )

    rows = m._run_all(args)

    assert seen_cfgs[0].input_len == 128
    assert seen_cfgs[0].random_prefix_len == 102
    assert rows[0]["input_len"] == 128
    assert rows[0]["total_input_len"] == 128
    assert rows[0]["prefix_tokens"] == 102
    assert rows[0]["prefix_ratio"] == 0.8
    assert "total_input=128 prefix=102tok(80%) suffix=26tok" in caplog.text


def test_run_all_rejects_invalid_prefix_ratio_before_serving(monkeypatch):
    async def fail_if_called(cfg):
        raise AssertionError("main_async should not be called")

    monkeypatch.setattr(serve, "main_async", fail_if_called)

    args = argparse.Namespace(
        model="m", served_model_name=None, backend="openai",
        base_url="http://x/v1", host="127.0.0.1", port=8000,
        insecure=False, api_key=None, tokenizer="/some/tok",
        input_lens=[128], output_lens=[8], cross_product=False,
        parallel_nums=[1], epochs=1, sleep_between=0,
        warmup_requests=0, prefix_ratio=float("nan"),
        seed=0, no_vary_seed_by_config=False,
        output_csv=None, output_xlsx=None,
        result_dir=None, max_ttft_ms=None, min_throughput_tok_s=None,
        min_output_compliance=0.0,
    )

    with pytest.raises(ValueError, match="--prefix-ratio"):
        m._run_all(args)


@pytest.mark.parametrize(
    "message",
    [
        "unique suffix capacity exhausted: random_input_len=2",
        "unique prompt collision: request_index=1",
    ],
)
def test_run_all_reraises_random_prompt_generation_errors(monkeypatch, message):
    async def fail_with_generation_error(cfg):
        raise ValueError(message)

    monkeypatch.setattr(serve, "main_async", fail_with_generation_error)

    args = argparse.Namespace(
        model="m", served_model_name=None, backend="openai",
        base_url="http://x/v1", host="127.0.0.1", port=8000,
        insecure=False, api_key=None, tokenizer="/some/tok",
        input_lens=[128], output_lens=[8], cross_product=False,
        parallel_nums=[1], epochs=1, sleep_between=0,
        warmup_requests=0, prefix_ratio=0.8,
        seed=0, no_vary_seed_by_config=False,
        output_csv=None, output_xlsx=None,
        result_dir=None, max_ttft_ms=None, min_throughput_tok_s=None,
        min_output_compliance=0.0,
    )

    with pytest.raises(ValueError, match=message.split(":")[0]):
        m._run_all(args)


def test_save_csv_persists_seed_column(tmp_path):
    csv_path = str(tmp_path / "bench.csv")

    m.save_csv([{"seed": 98765}], csv_path)

    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        data = list(reader)

    assert "seed" in (reader.fieldnames or [])
    assert data[0]["seed"] == "98765"


def test_seed_column_is_reported_after_num_prompts():
    seed_index = m.CSV_HEADERS.index("seed")

    assert m.CSV_HEADERS[m.CSV_HEADERS.index("num_prompts") + 1] == "seed"
    assert m.CSV_HEADERS_ZH[seed_index] == "随机种子"
    assert len(m.CSV_HEADERS) == len(m.CSV_HEADERS_ZH)


def test_save_xlsx_persists_seed_column(tmp_path):
    pytest.importorskip("openpyxl")
    xlsx_path = str(tmp_path / "bench.xlsx")

    m.save_xlsx([{"seed": 98765}], xlsx_path)

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["基准测试结果"]
    headers = [cell.value for cell in ws[1]]
    seed_col = headers.index("seed") + 1

    assert headers[headers.index("num_prompts") + 1] == "seed"
    assert ws.cell(row=3, column=seed_col).value == 98765


def _run_and_capture_seeds(monkeypatch, *, seed=123, no_vary_seed_by_config=False):
    import argparse

    captured = []

    async def _fake_main_async(cfg):
        captured.append(cfg.seed)
        return _fake_result(cfg.input_len, cfg.output_len, completed=cfg.num_prompts)

    monkeypatch.setattr(serve, "main_async", _fake_main_async)

    our_args = argparse.Namespace(
        model="m", served_model_name=None, backend="openai-chat",
        base_url=None, host="127.0.0.1", port=8000, insecure=False, api_key=None,
        tokenizer="/some/tok",
        input_lens=[128], output_lens=[8], cross_product=False,
        parallel_nums=[1, 4, 8], epochs=1, sleep_between=0, warmup_requests=0,
        prefix_ratio=0.8,
        seed=seed, no_vary_seed_by_config=no_vary_seed_by_config,
        max_ttft_ms=None, min_throughput_tok_s=None, min_output_compliance=0.95,
        output_csv=None, output_xlsx=None, result_dir=None,
    )
    rows = m._run_all(our_args)
    return captured, rows


def test_run_all_rejects_invalid_seed_before_benchmark(monkeypatch):
    import argparse

    async def _fake_main_async(cfg):
        pytest.fail("main_async should not be called for an invalid seed")

    monkeypatch.setattr(serve, "main_async", _fake_main_async)

    our_args = argparse.Namespace(
        model="m", served_model_name=None, backend="openai-chat",
        base_url=None, host="127.0.0.1", port=8000, insecure=False, api_key=None,
        tokenizer="/some/tok",
        input_lens=[128], output_lens=[8], cross_product=False,
        parallel_nums=[1], epochs=1, sleep_between=0, warmup_requests=0,
        prefix_ratio=0.0,
        seed=2**32, no_vary_seed_by_config=False,
        max_ttft_ms=None, min_throughput_tok_s=None, min_output_compliance=0.95,
        output_csv=None, output_xlsx=None, result_dir=None,
    )

    with pytest.raises(ValueError, match="--seed"):
        m._run_all(our_args)


def test_run_all_varies_seed_by_config_by_default(monkeypatch):
    captured, rows = _run_and_capture_seeds(monkeypatch)
    expected = [
        m.derive_config_seed(
            base_seed=123,
            input_len=128,
            output_len=8,
            parallel_num=parallel_num,
            prefix_ratio=0.8,
            config_index=config_index,
        )
        for config_index, parallel_num in enumerate([1, 4, 8], start=1)
    ]

    assert captured == expected
    assert [row["seed"] for row in rows] == expected


def test_run_all_uses_current_io_pair_when_deriving_seed(monkeypatch):
    import argparse

    captured = []

    async def _fake_main_async(cfg):
        captured.append((cfg.input_len, cfg.output_len, cfg.max_concurrency, cfg.seed))
        return _fake_result(cfg.input_len, cfg.output_len, completed=cfg.num_prompts)

    monkeypatch.setattr(serve, "main_async", _fake_main_async)

    our_args = argparse.Namespace(
        model="m", served_model_name=None, backend="openai-chat",
        base_url=None, host="127.0.0.1", port=8000, insecure=False, api_key=None,
        tokenizer="/some/tok",
        input_lens=[128, 256], output_lens=[8, 16], cross_product=False,
        parallel_nums=[1], epochs=1, sleep_between=0, warmup_requests=0,
        prefix_ratio=0.25,
        seed=123, no_vary_seed_by_config=False,
        max_ttft_ms=None, min_throughput_tok_s=None, min_output_compliance=0.95,
        output_csv=None, output_xlsx=None, result_dir=None,
    )
    rows = m._run_all(our_args)

    expected_configs = [(128, 8, 1), (256, 16, 1)]
    expected_seeds = [
        m.derive_config_seed(
            base_seed=123,
            input_len=input_len,
            output_len=output_len,
            parallel_num=parallel_num,
            prefix_ratio=0.25,
            config_index=config_index,
        )
        for config_index, (input_len, output_len, parallel_num)
        in enumerate(expected_configs, start=1)
    ]

    assert captured == [
        (input_len, output_len, parallel_num, seed)
        for (input_len, output_len, parallel_num), seed
        in zip(expected_configs, expected_seeds)
    ]
    assert [row["seed"] for row in rows] == expected_seeds


def test_run_all_default_seed_derivation_depends_on_base_seed(monkeypatch):
    first_captured, _ = _run_and_capture_seeds(monkeypatch, seed=123)
    second_captured, _ = _run_and_capture_seeds(monkeypatch, seed=456)

    assert first_captured != second_captured


def test_run_all_can_use_fixed_seed_for_compatibility(monkeypatch):
    captured, rows = _run_and_capture_seeds(
        monkeypatch,
        no_vary_seed_by_config=True,
    )

    assert captured == [123, 123, 123]
    assert [row["seed"] for row in rows] == [123, 123, 123]
