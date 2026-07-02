import csv
import json
import threading
from pathlib import Path

import pytest

import resource_monitor as rm


def test_nvidia_smi_query_is_subprocess_argv():
    assert rm.NVIDIA_SMI_QUERY == [
        "nvidia-smi",
        "--query-gpu=index,name,uuid,utilization.gpu,memory.used,memory.total,power.draw,temperature.gpu",
        "--format=csv,noheader,nounits",
    ]


def test_default_readers_nvidia_smi_uses_timeout(monkeypatch):
    calls = []

    class Completed:
        stdout = "0, NVIDIA H200, GPU-0, 50, 1000, 2000, 300, 70\n"

    def fake_run(argv, **kwargs):
        calls.append((argv, kwargs))
        return Completed()

    monkeypatch.setattr(rm.subprocess, "run", fake_run)

    assert rm.default_readers().nvidia_smi() == Completed.stdout
    assert calls == [(
        rm.NVIDIA_SMI_QUERY,
        {
            "check": True,
            "capture_output": True,
            "text": True,
            "timeout": 10,
        },
    )]


def test_parse_cpu_stat_and_compute_utilization():
    first = rm.parse_proc_stat("cpu  100 0 50 850 0 0 0 0 0 0\n")
    second = rm.parse_proc_stat("cpu  180 0 70 950 0 0 0 0 0 0\n")

    assert rm.cpu_utilization_pct(first, second) == pytest.approx(50.0)
    assert rm.cpu_utilization_pct(first, first) is None
    assert rm.cpu_utilization_pct(second, first) is None


def test_parse_meminfo_uses_mem_available():
    sample = "\n".join([
        "MemTotal:       1024000 kB",
        "MemFree:         100000 kB",
        "MemAvailable:    256000 kB",
    ])

    memory = rm.parse_meminfo(sample)

    assert memory["mem_total_mb"] == pytest.approx(1000.0)
    assert memory["mem_available_mb"] == pytest.approx(250.0)
    assert memory["mem_used_mb"] == pytest.approx(750.0)
    assert memory["mem_used_pct"] == pytest.approx(75.0)


def test_parse_net_dev_skips_loopback_and_computes_rates():
    before = """
Inter-|   Receive                                                |  Transmit
 face |bytes    packets errs drop fifo frame compressed multicast|bytes    packets errs drop fifo colls carrier compressed
    lo: 1000 0 0 0 0 0 0 0 2000 0 0 0 0 0 0 0
  eth0: 1048576 0 0 0 0 0 0 0 2097152 0 0 0 0 0 0 0
"""
    after = """
Inter-|   Receive                                                |  Transmit
 face |bytes    packets errs drop fifo frame compressed multicast|bytes    packets errs drop fifo colls carrier compressed
    lo: 999999 0 0 0 0 0 0 0 999999 0 0 0 0 0 0 0
  eth0: 3145728 0 0 0 0 0 0 0 5242880 0 0 0 0 0 0 0
"""

    rates = rm.network_rates_mb_s(
        rm.parse_net_dev(before),
        rm.parse_net_dev(after),
        elapsed_s=2.0,
    )

    assert rates["net_rx_mb_s"] == pytest.approx(1.0)
    assert rates["net_tx_mb_s"] == pytest.approx(1.5)

    invalid_rates = rm.network_rates_mb_s(
        rm.parse_net_dev(before),
        rm.parse_net_dev(after),
        elapsed_s=0.0,
    )
    assert invalid_rates["net_rx_mb_s"] is None
    assert invalid_rates["net_tx_mb_s"] is None


def test_parse_nvidia_smi_csv_handles_multi_gpu_and_empty_values():
    text = "\n".join([
        "0, NVIDIA H200, GPU-0, 98, 77320, 81559, 612.50, 76",
        "1, NVIDIA H200, GPU-1, N/A, 12000, 81559, Not Supported, Not Supported",
        "2, NVIDIA H200, GPU-2, , 0, 81559, [Not Supported], ",
    ])

    gpus = rm.parse_nvidia_smi_csv(text)

    assert gpus[0]["gpu_index"] == 0
    assert gpus[0]["gpu_util_pct"] == 98.0
    assert gpus[0]["gpu_mem_used_pct"] == pytest.approx(94.8001, rel=0.001)
    assert gpus[1]["gpu_util_pct"] is None
    assert gpus[1]["gpu_power_w"] is None
    assert gpus[1]["gpu_temperature_c"] is None
    assert gpus[2]["gpu_index"] == 2
    assert gpus[2]["gpu_util_pct"] is None
    assert gpus[2]["gpu_power_w"] is None
    assert gpus[2]["gpu_temperature_c"] is None


def test_summarize_samples_computes_avg_p95_and_max():
    samples = [
        {
            "cpu_util_pct": 10.0,
            "mem_used_mb": 100.0,
            "mem_used_pct": 25.0,
            "net_rx_mb_s": 1.0,
            "net_tx_mb_s": 0.5,
            "disk_read_mb_s": 10.0,
            "disk_write_mb_s": 4.0,
            "gpu_util_avg_pct": 20.0,
            "gpu_mem_used_mb": 1000.0,
            "gpu_mem_total_mb": 8000.0,
            "gpu_mem_used_pct": 12.5,
            "gpu_power_w": 100.0,
            "gpu_temperature_c": 50.0,
        },
        {
            "cpu_util_pct": 20.0,
            "mem_used_mb": 200.0,
            "mem_used_pct": 50.0,
            "net_rx_mb_s": 2.0,
            "net_tx_mb_s": 1.0,
            "disk_read_mb_s": 20.0,
            "disk_write_mb_s": 8.0,
            "gpu_util_avg_pct": 40.0,
            "gpu_mem_used_mb": 2000.0,
            "gpu_mem_total_mb": 8000.0,
            "gpu_mem_used_pct": 25.0,
            "gpu_power_w": 200.0,
            "gpu_temperature_c": 60.0,
        },
        {
            "cpu_util_pct": 30.0,
            "mem_used_mb": 300.0,
            "mem_used_pct": 75.0,
            "net_rx_mb_s": 3.0,
            "net_tx_mb_s": 1.5,
            "disk_read_mb_s": 30.0,
            "disk_write_mb_s": 12.0,
            "gpu_util_avg_pct": 60.0,
            "gpu_mem_used_mb": 3000.0,
            "gpu_mem_total_mb": 8000.0,
            "gpu_mem_used_pct": 37.5,
            "gpu_power_w": 300.0,
            "gpu_temperature_c": 70.0,
        },
    ]
    gpu_details = [{"gpu_index": 0}, {"gpu_index": 1}]

    summary = rm.summarize_samples(samples, gpu_details=gpu_details)

    assert summary["available"] is True
    assert summary["system_available"] is True
    assert summary["gpu_available"] is True
    assert summary["gpus"] == gpu_details
    assert "gpu_details" not in summary
    assert summary["sample_count"] == 3

    assert set(summary["aggregate"]) == {
        "cpu_util_avg_pct",
        "cpu_util_p95_pct",
        "cpu_util_max_pct",
        "mem_used_avg_mb",
        "mem_used_p95_mb",
        "mem_used_max_mb",
        "mem_used_avg_pct",
        "mem_used_p95_pct",
        "mem_used_max_pct",
        "net_rx_avg_mb_s",
        "net_rx_p95_mb_s",
        "net_rx_max_mb_s",
        "net_tx_avg_mb_s",
        "net_tx_p95_mb_s",
        "net_tx_max_mb_s",
        "disk_read_avg_mb_s",
        "disk_read_p95_mb_s",
        "disk_read_max_mb_s",
        "disk_write_avg_mb_s",
        "disk_write_p95_mb_s",
        "disk_write_max_mb_s",
        "gpu_count",
        "gpu_util_avg_pct",
        "gpu_util_p95_pct",
        "gpu_util_max_pct",
        "gpu_mem_used_avg_mb",
        "gpu_mem_used_p95_mb",
        "gpu_mem_used_max_mb",
        "gpu_mem_total_mb",
        "gpu_mem_used_max_pct",
        "gpu_power_avg_w",
        "gpu_power_p95_w",
        "gpu_power_max_w",
        "gpu_temp_avg_c",
        "gpu_temp_p95_c",
        "gpu_temp_max_c",
    }
    aggregate = summary["aggregate"]
    assert aggregate["cpu_util_avg_pct"] == pytest.approx(20.0)
    assert aggregate["cpu_util_p95_pct"] == pytest.approx(30.0)
    assert aggregate["cpu_util_max_pct"] == pytest.approx(30.0)
    assert aggregate["mem_used_avg_mb"] == pytest.approx(200.0)
    assert aggregate["mem_used_p95_mb"] == pytest.approx(300.0)
    assert aggregate["mem_used_max_mb"] == pytest.approx(300.0)
    assert aggregate["mem_used_avg_pct"] == pytest.approx(50.0)
    assert aggregate["mem_used_p95_pct"] == pytest.approx(75.0)
    assert aggregate["mem_used_max_pct"] == pytest.approx(75.0)
    assert aggregate["net_rx_avg_mb_s"] == pytest.approx(2.0)
    assert aggregate["net_rx_p95_mb_s"] == pytest.approx(3.0)
    assert aggregate["net_rx_max_mb_s"] == pytest.approx(3.0)
    assert aggregate["net_tx_avg_mb_s"] == pytest.approx(1.0)
    assert aggregate["net_tx_p95_mb_s"] == pytest.approx(1.5)
    assert aggregate["net_tx_max_mb_s"] == pytest.approx(1.5)
    assert aggregate["disk_read_avg_mb_s"] == pytest.approx(20.0)
    assert aggregate["disk_read_p95_mb_s"] == pytest.approx(30.0)
    assert aggregate["disk_read_max_mb_s"] == pytest.approx(30.0)
    assert aggregate["disk_write_avg_mb_s"] == pytest.approx(8.0)
    assert aggregate["disk_write_p95_mb_s"] == pytest.approx(12.0)
    assert aggregate["disk_write_max_mb_s"] == pytest.approx(12.0)
    assert aggregate["gpu_count"] == 2
    assert aggregate["gpu_util_avg_pct"] == pytest.approx(40.0)
    assert aggregate["gpu_util_p95_pct"] == pytest.approx(60.0)
    assert aggregate["gpu_util_max_pct"] == pytest.approx(60.0)
    assert aggregate["gpu_mem_used_avg_mb"] == pytest.approx(2000.0)
    assert aggregate["gpu_mem_used_p95_mb"] == pytest.approx(3000.0)
    assert aggregate["gpu_mem_used_max_mb"] == pytest.approx(3000.0)
    assert aggregate["gpu_mem_total_mb"] == pytest.approx(8000.0)
    assert aggregate["gpu_mem_used_max_pct"] == pytest.approx(37.5)
    assert aggregate["gpu_power_avg_w"] == pytest.approx(200.0)
    assert aggregate["gpu_power_p95_w"] == pytest.approx(300.0)
    assert aggregate["gpu_power_max_w"] == pytest.approx(300.0)
    assert aggregate["gpu_temp_avg_c"] == pytest.approx(60.0)
    assert aggregate["gpu_temp_p95_c"] == pytest.approx(70.0)
    assert aggregate["gpu_temp_max_c"] == pytest.approx(70.0)


def test_summarize_samples_reports_unavailable_when_empty():
    summary = rm.summarize_samples([], gpu_details=[])

    assert summary["available"] is False
    assert summary["system_available"] is False
    assert summary["gpu_available"] is False
    assert summary["sample_count"] == 0


def test_summarize_samples_uses_case_level_gpu_sample_fields():
    samples = [
        {
            "gpu_util_avg_pct": 30.0,
            "gpu_mem_used_avg_mb": 1000.0,
            "gpu_mem_used_max_mb": 1500.0,
            "gpu_mem_total_mb": 8000.0,
            "gpu_mem_used_max_pct": 18.75,
            "gpu_power_avg_w": 200.0,
            "gpu_power_max_w": 250.0,
            "gpu_temp_max_c": 60.0,
        },
        {
            "gpu_util_avg_pct": 60.0,
            "gpu_mem_used_avg_mb": 2000.0,
            "gpu_mem_used_max_mb": 2500.0,
            "gpu_mem_total_mb": 8000.0,
            "gpu_mem_used_max_pct": 31.25,
            "gpu_power_avg_w": 300.0,
            "gpu_power_max_w": 350.0,
            "gpu_temp_max_c": 70.0,
        },
    ]

    summary = rm.summarize_samples(samples, gpu_details=[{"gpu_index": 0}])

    aggregate = summary["aggregate"]
    assert aggregate["gpu_mem_used_avg_mb"] == pytest.approx(1500.0)
    assert aggregate["gpu_mem_used_p95_mb"] == pytest.approx(2000.0)
    assert aggregate["gpu_mem_used_max_mb"] == pytest.approx(2500.0)
    assert aggregate["gpu_mem_used_max_pct"] == pytest.approx(31.25)
    assert aggregate["gpu_power_avg_w"] == pytest.approx(250.0)
    assert aggregate["gpu_power_p95_w"] == pytest.approx(300.0)
    assert aggregate["gpu_power_max_w"] == pytest.approx(350.0)


def test_resource_monitor_writes_samples_and_summary(tmp_path):
    monitor = rm.ResourceMonitor(
        output_dir=tmp_path,
        interval_sec=1.0,
        enabled=True,
        backend="nvidia-smi",
        readers=rm.ResourceReaders(
            proc_stat=lambda: "cpu  100 0 50 850 0 0 0 0 0 0\n",
            meminfo=lambda: "MemTotal: 1024000 kB\nMemAvailable: 512000 kB\n",
            net_dev=lambda: "",
            diskstats=lambda: "",
            nvidia_smi=lambda: "0, NVIDIA H200, GPU-0, 50, 1000, 2000, 300, 70\n",
        ),
    )

    monitor.sample_once(now=100.0)
    summary = monitor.stop()

    assert (tmp_path / "resource_samples.csv").is_file()
    assert (tmp_path / "resource_summary.json").is_file()
    loaded = json.loads((tmp_path / "resource_summary.json").read_text(encoding="utf-8"))
    assert loaded["available"] is True
    assert loaded["gpu_available"] is True
    assert summary["aggregate"]["gpu_count"] == 1


def test_resource_monitor_stop_writes_samples_from_snapshot(tmp_path, monkeypatch):
    monitor = rm.ResourceMonitor(
        output_dir=tmp_path,
        interval_sec=1.0,
        enabled=True,
        backend="nvidia-smi",
        readers=rm.ResourceReaders(
            proc_stat=lambda: "cpu  100 0 50 850 0 0 0 0 0 0\n",
            meminfo=lambda: "MemTotal: 1024000 kB\nMemAvailable: 512000 kB\n",
            net_dev=lambda: "",
            diskstats=lambda: "",
            nvidia_smi=lambda: "0, NVIDIA H200, GPU-0, 50, 1000, 2000, 300, 70\n",
        ),
    )
    captured = {}

    def fake_write_samples_csv(path, samples):
        monitor.samples.append({
            "timestamp": "late",
            "elapsed_s": 999.0,
            "mem_total_mb": 1.0,
            "mem_used_mb": 1.0,
            "mem_available_mb": 0.0,
            "mem_used_pct": 100.0,
        })
        captured["path"] = Path(path)
        captured["samples"] = list(samples)

    monkeypatch.setattr(rm, "write_samples_csv", fake_write_samples_csv)

    monitor.sample_once(now=100.0)
    summary = monitor.stop()

    assert summary["sample_count"] == 1
    assert captured["path"] == tmp_path / "resource_samples.csv"
    assert len(captured["samples"]) == summary["sample_count"]
    assert captured["samples"][0]["timestamp"] != "late"


def test_resource_monitor_stop_takes_final_sample_for_short_lifecycle(tmp_path, monkeypatch):
    times = iter([100.0, 101.0])

    def fake_time():
        return next(times)

    def sequence(values):
        values = iter(values)
        last = None

        def read():
            nonlocal last
            try:
                last = next(values)
            except StopIteration:
                pass
            return last

        return read

    proc_stat = sequence([
        "cpu  100 0 0 900 0 0 0 0 0 0\n",
        "cpu  150 0 0 950 0 0 0 0 0 0\n",
    ])
    net_dev = sequence([
        """
Inter-|   Receive                                                |  Transmit
 face |bytes    packets errs drop fifo frame compressed multicast|bytes    packets errs drop fifo colls carrier compressed
  eth0: 1048576 0 0 0 0 0 0 0 2097152 0 0 0 0 0 0 0
""",
        """
Inter-|   Receive                                                |  Transmit
 face |bytes    packets errs drop fifo frame compressed multicast|bytes    packets errs drop fifo colls carrier compressed
  eth0: 3145728 0 0 0 0 0 0 0 5242880 0 0 0 0 0 0 0
""",
    ])
    diskstats = sequence([
        "   8       0 sda 10 0 100 0 5 0 20 0 0 0 0 0 0 0 0 0 0\n",
        "   8       0 sda 20 0 2148 0 7 0 1044 0 0 0 0 0 0 0 0 0 0\n",
    ])
    monitor = rm.ResourceMonitor(
        output_dir=tmp_path,
        interval_sec=3600.0,
        enabled=True,
        backend="nvidia-smi",
        readers=rm.ResourceReaders(
            proc_stat=proc_stat,
            meminfo=lambda: "MemTotal: 1024000 kB\nMemAvailable: 512000 kB\n",
            net_dev=net_dev,
            diskstats=diskstats,
            nvidia_smi=lambda: "",
        ),
    )
    monkeypatch.setattr(rm.time, "time", fake_time)

    monitor.start()
    summary = monitor.stop()

    assert summary["sample_count"] >= 2
    aggregate = summary["aggregate"]
    for key in (
        "cpu_util_avg_pct",
        "net_rx_avg_mb_s",
        "net_tx_avg_mb_s",
        "disk_read_avg_mb_s",
        "disk_write_avg_mb_s",
    ):
        assert aggregate[key] is not None

    samples_path = tmp_path / "resource_samples.csv"
    summary_path = tmp_path / "resource_summary.json"
    assert samples_path.is_file()
    assert summary_path.is_file()
    with samples_path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert len(rows) >= 2


def test_resource_monitor_degrades_when_nvidia_smi_missing(tmp_path):
    def missing_gpu():
        raise FileNotFoundError("nvidia-smi")

    monitor = rm.ResourceMonitor(
        output_dir=tmp_path,
        interval_sec=1.0,
        enabled=True,
        backend="nvidia-smi",
        readers=rm.ResourceReaders(
            proc_stat=lambda: "cpu  100 0 50 850 0 0 0 0 0 0\n",
            meminfo=lambda: "MemTotal: 1024000 kB\nMemAvailable: 512000 kB\n",
            net_dev=lambda: "",
            diskstats=lambda: "",
            nvidia_smi=missing_gpu,
        ),
    )

    monitor.sample_once(now=100.0)
    summary = monitor.stop()

    assert summary["available"] is True
    assert summary["system_available"] is True
    assert summary["gpu_available"] is False
    assert summary["error_count"] >= 1


def test_resource_monitor_reports_unavailable_when_all_readers_fail(tmp_path):
    def fail():
        raise OSError("unavailable")

    monitor = rm.ResourceMonitor(
        output_dir=tmp_path,
        interval_sec=1.0,
        enabled=True,
        backend="nvidia-smi",
        readers=rm.ResourceReaders(
            proc_stat=fail,
            meminfo=fail,
            net_dev=fail,
            diskstats=fail,
            nvidia_smi=fail,
        ),
    )

    monitor.sample_once(now=100.0)
    summary = monitor.stop()

    assert summary["available"] is False
    assert summary["system_available"] is False
    assert summary["gpu_available"] is False
    assert summary["sample_count"] == 1
    assert summary["error_count"] > 0


@pytest.mark.parametrize(
    "failing_reader",
    ["proc_stat", "meminfo", "net_dev", "diskstats", "nvidia_smi"],
)
def test_resource_monitor_passthrough_exceptions_are_not_swallowed(
    tmp_path,
    failing_reader,
):
    class PassthroughError(Exception):
        pass

    def fail():
        raise PassthroughError("stop requested")

    reader_values = {
        "proc_stat": lambda: "cpu  100 0 50 850 0 0 0 0 0 0\n",
        "meminfo": lambda: "MemTotal: 1024000 kB\nMemAvailable: 512000 kB\n",
        "net_dev": lambda: "",
        "diskstats": lambda: "",
        "nvidia_smi": lambda: "",
    }
    reader_values[failing_reader] = fail
    readers = rm.ResourceReaders(**reader_values)
    degraded_monitor = rm.ResourceMonitor(
        output_dir=tmp_path / "degraded",
        interval_sec=1.0,
        enabled=True,
        backend="nvidia-smi",
        readers=readers,
    )

    degraded_monitor.sample_once(now=100.0)

    assert degraded_monitor.error_count >= 1

    passthrough_monitor = rm.ResourceMonitor(
        output_dir=tmp_path / "passthrough",
        interval_sec=1.0,
        enabled=True,
        backend="nvidia-smi",
        readers=readers,
        passthrough_exceptions=(PassthroughError,),
    )

    with pytest.raises(PassthroughError, match="stop requested"):
        passthrough_monitor.sample_once(now=100.0)


def test_resource_monitor_stop_reraises_background_passthrough_exception(tmp_path):
    class PassthroughError(Exception):
        pass

    sample_attempts = 0
    background_sample_started = threading.Event()

    def proc_stat():
        nonlocal sample_attempts
        sample_attempts += 1
        if sample_attempts == 2:
            background_sample_started.set()
            raise PassthroughError("background stop requested")
        return "cpu  100 0 50 850 0 0 0 0 0 0\n"

    monitor = rm.ResourceMonitor(
        output_dir=tmp_path,
        interval_sec=0.01,
        enabled=True,
        backend="nvidia-smi",
        readers=rm.ResourceReaders(
            proc_stat=proc_stat,
            meminfo=lambda: "MemTotal: 1024000 kB\nMemAvailable: 512000 kB\n",
            net_dev=lambda: "",
            diskstats=lambda: "",
            nvidia_smi=lambda: "",
        ),
        passthrough_exceptions=(PassthroughError,),
    )

    monitor.start()
    assert background_sample_started.wait(timeout=1.0)

    with pytest.raises(PassthroughError, match="background stop requested"):
        monitor.stop()


def test_resource_monitor_stop_reraises_late_background_passthrough_exception(
    tmp_path,
    monkeypatch,
):
    class PassthroughError(Exception):
        pass

    sample_attempts = 0
    background_reader_entered = threading.Event()
    allow_background_error = threading.Event()

    def proc_stat():
        nonlocal sample_attempts
        sample_attempts += 1
        if sample_attempts == 2:
            background_reader_entered.set()
            assert allow_background_error.wait(timeout=1.0)
            raise PassthroughError("late background stop requested")
        return "cpu  100 0 50 850 0 0 0 0 0 0\n"

    monitor = rm.ResourceMonitor(
        output_dir=tmp_path,
        interval_sec=0.01,
        enabled=True,
        backend="nvidia-smi",
        readers=rm.ResourceReaders(
            proc_stat=proc_stat,
            meminfo=lambda: "MemTotal: 1024000 kB\nMemAvailable: 512000 kB\n",
            net_dev=lambda: "",
            diskstats=lambda: "",
            nvidia_smi=lambda: "",
        ),
        passthrough_exceptions=(PassthroughError,),
    )

    monitor.start()
    assert background_reader_entered.wait(timeout=1.0)

    original_join = monitor._thread.join
    join_calls = 0

    def fake_join(timeout=None):
        nonlocal join_calls
        join_calls += 1
        if join_calls == 1:
            return
        allow_background_error.set()
        original_join(timeout=1.0)

    original_sample_once = monitor.sample_once

    def sample_once(*args, **kwargs):
        if threading.current_thread() is not monitor._thread:
            allow_background_error.set()
        return original_sample_once(*args, **kwargs)

    monkeypatch.setattr(monitor._thread, "join", fake_join)
    monkeypatch.setattr(monitor, "sample_once", sample_once)

    with pytest.raises(PassthroughError, match="late background stop requested"):
        monitor.stop()

    assert join_calls >= 2


def test_parse_diskstats_and_compute_rates():
    before = "   8       0 sda 10 0 100 0 5 0 20 0 0 0 0 0 0 0 0 0 0\n"
    after = "   8       0 sda 20 0 4196 0 7 0 2068 0 0 0 0 0 0 0 0 0 0\n"

    rates = rm.disk_rates_mb_s(
        rm.parse_diskstats(before),
        rm.parse_diskstats(after),
        elapsed_s=2.0,
    )

    assert rates["disk_read_mb_s"] == pytest.approx(1.0)
    assert rates["disk_write_mb_s"] == pytest.approx(0.5)


def test_parse_diskstats_skips_partitions_and_mapped_devices():
    text = "\n".join([
        "   8       0 sda 0 0 100 0 0 0 50 0 0 0 0 0 0 0 0 0",
        "   8       1 sda1 0 0 10000 0 0 0 5000 0 0 0 0 0 0 0 0 0",
        " 253       0 dm-0 0 0 20000 0 0 0 10000 0 0 0 0 0 0 0 0 0",
        "   9       0 md0 0 0 40000 0 0 0 20000 0 0 0 0 0 0 0 0 0",
    ])

    stats = rm.parse_diskstats(text)

    assert stats["read_sectors"] == 100
    assert stats["write_sectors"] == 50
    assert set(stats["devices"]) == {"sda"}


def test_append_summary_to_result_csv_adds_resource_columns(tmp_path):
    result_csv = tmp_path / "result.csv"
    result_csv.write_text("model,throughput_tok_s\nm,12.5\n", encoding="utf-8-sig")
    summary = {
        "available": True,
        "sample_count": 2,
        "aggregate": {
            "cpu_util_avg_pct": 50.0,
            "gpu_count": 1,
            "gpu_util_max_pct": 99.0,
        },
    }

    rm.append_summary_to_result_files(tmp_path, summary)

    with result_csv.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert rows[0]["model"] == "m"
    assert rows[0]["resource_monitor_available"] == "true"
    assert rows[0]["resource_sample_count"] == "2"
    assert rows[0]["cpu_util_avg_pct"] == "50.0"
    assert rows[0]["gpu_util_max_pct"] == "99.0"


def test_append_prefixed_summaries_to_result_csv(tmp_path):
    result_csv = tmp_path / "result.csv"
    result_csv.write_text("model,throughput_tok_s\nm,12.5\n", encoding="utf-8-sig")
    summary = {
        "available": True,
        "sample_count": 2,
        "aggregate": {"cpu_util_avg_pct": 50.0, "gpu_mem_used_max_mb": 1234.0},
    }

    rm.append_prefixed_summaries_to_result_files(
        tmp_path,
        {"p1": summary, "router": summary},
    )

    with result_csv.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert rows[0]["p1_resource_monitor_available"] == "true"
    assert rows[0]["p1_cpu_util_avg_pct"] == "50.0"
    assert rows[0]["router_gpu_mem_used_max_mb"] == "1234.0"


def test_append_prefixed_summaries_to_result_xlsx(tmp_path):
    import openpyxl

    result_xlsx = tmp_path / "result.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.cell(row=1, column=1, value="model")
    worksheet.cell(row=1, column=2, value="throughput_tok_s")
    worksheet.cell(row=2, column=1, value="模型")
    worksheet.cell(row=2, column=2, value="吞吐")
    worksheet.cell(row=3, column=1, value="m")
    worksheet.cell(row=3, column=2, value=12.5)
    workbook.save(result_xlsx)
    summary = {
        "available": True,
        "sample_count": 2,
        "aggregate": {"cpu_util_avg_pct": 50.0, "gpu_mem_used_max_mb": 1234.0},
    }

    rm.append_prefixed_summaries_to_result_files(
        tmp_path,
        {"p1": summary, "router": summary},
    )

    loaded = openpyxl.load_workbook(result_xlsx)
    sheet = loaded.active
    columns = {
        sheet.cell(row=1, column=column).value: column
        for column in range(1, sheet.max_column + 1)
    }
    assert sheet.cell(row=1, column=columns["model"]).value == "model"
    assert sheet.cell(row=2, column=columns["model"]).value == "模型"
    assert sheet.cell(row=3, column=columns["model"]).value == "m"
    assert sheet.cell(row=3, column=columns["throughput_tok_s"]).value == 12.5

    p1_available = columns["p1_resource_monitor_available"]
    p1_cpu_avg = columns["p1_cpu_util_avg_pct"]
    router_gpu_mem = columns["router_gpu_mem_used_max_mb"]
    assert sheet.cell(row=2, column=p1_available).value == "p1_resource_monitor_available"
    assert sheet.cell(row=2, column=p1_cpu_avg).value == "p1_cpu_util_avg_pct"
    assert sheet.cell(row=2, column=router_gpu_mem).value == "router_gpu_mem_used_max_mb"
    assert sheet.cell(row=3, column=p1_available).value == "true"
    assert sheet.cell(row=3, column=p1_cpu_avg).value == 50.0
    assert sheet.cell(row=3, column=router_gpu_mem).value == 1234.0


def test_append_summary_to_result_xlsx_adds_chinese_resource_headers(tmp_path):
    import openpyxl

    result_xlsx = tmp_path / "result.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.cell(row=1, column=1, value="model")
    worksheet.cell(row=1, column=2, value="throughput_tok_s")
    worksheet.cell(row=2, column=1, value="模型")
    worksheet.cell(row=2, column=2, value="吞吐")
    worksheet.cell(row=3, column=1, value="m")
    worksheet.cell(row=3, column=2, value=12.5)
    workbook.save(result_xlsx)

    summary = {
        "available": True,
        "sample_count": 2,
        "aggregate": {
            "cpu_util_avg_pct": 50.0,
            "gpu_count": 1,
            "gpu_util_max_pct": 99.0,
        },
    }

    rm.append_summary_to_result_files(tmp_path, summary)

    loaded = openpyxl.load_workbook(result_xlsx)
    sheet = loaded.active
    columns = {
        sheet.cell(row=1, column=column).value: column
        for column in range(1, sheet.max_column + 1)
    }
    available_column = columns["resource_monitor_available"]
    sample_count_column = columns["resource_sample_count"]
    cpu_avg_column = columns["cpu_util_avg_pct"]

    assert sheet.cell(row=2, column=available_column).value == "资源监控可用"
    assert sheet.cell(row=2, column=sample_count_column).value == "资源采样数"
    assert sheet.cell(row=2, column=cpu_avg_column).value == "CPU平均使用率(%)"
    for column_name in rm.RESOURCE_RESULT_COLUMNS:
        column = columns[column_name]
        assert sheet.cell(row=2, column=column).value
        assert sheet.cell(row=2, column=column).value != column_name
    assert sheet.cell(row=3, column=available_column).value == "true"
    assert sheet.cell(row=3, column=sample_count_column).value == 2
    assert sheet.cell(row=3, column=cpu_avg_column).value == 50.0


def test_append_summary_to_result_files_propagates_xlsx_write_errors(tmp_path, monkeypatch):
    (tmp_path / "result.xlsx").write_bytes(b"not a workbook")

    def fail_xlsx(path, values):
        raise RuntimeError("xlsx failed")

    monkeypatch.setattr(rm, "append_summary_to_xlsx", fail_xlsx)

    with pytest.raises(RuntimeError, match="xlsx failed"):
        rm.append_summary_to_result_files(
            tmp_path,
            {"available": True, "sample_count": 1, "aggregate": {}},
        )


def test_resource_monitor_disabled_writes_nothing(tmp_path):
    monitor = rm.ResourceMonitor(
        output_dir=tmp_path,
        interval_sec=1.0,
        enabled=False,
        backend="nvidia-smi",
    )

    monitor.start()
    monitor.sample_once(now=100.0)
    summary = monitor.stop()

    assert summary["available"] is False
    assert summary["sample_count"] == 0
    assert not (tmp_path / "resource_samples.csv").exists()
    assert not (tmp_path / "resource_summary.json").exists()
