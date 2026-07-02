import csv
import json
import math
import subprocess
import threading
import time
from dataclasses import dataclass
from io import StringIO
from pathlib import Path
from typing import Any, Callable, Mapping


NVIDIA_SMI_QUERY = [
    "nvidia-smi",
    "--query-gpu=index,name,uuid,utilization.gpu,memory.used,memory.total,power.draw,temperature.gpu",
    "--format=csv,noheader,nounits",
]

_MB = 1024.0 * 1024.0
_SECTOR_BYTES = 512.0

SAMPLE_HEADERS = [
    "timestamp", "elapsed_s",
    "cpu_util_pct",
    "mem_total_mb", "mem_used_mb", "mem_available_mb", "mem_used_pct",
    "net_rx_mb_s", "net_tx_mb_s",
    "disk_read_mb_s", "disk_write_mb_s",
    "gpu_count", "gpu_util_avg_pct", "gpu_util_max_pct",
    "gpu_mem_used_avg_mb", "gpu_mem_used_max_mb", "gpu_mem_total_mb",
    "gpu_mem_used_max_pct", "gpu_power_avg_w", "gpu_power_max_w",
    "gpu_temp_max_c",
]

RESOURCE_RESULT_COLUMNS = [
    "resource_monitor_available", "resource_sample_count",
    "cpu_util_avg_pct", "cpu_util_p95_pct", "cpu_util_max_pct",
    "mem_used_avg_mb", "mem_used_p95_mb", "mem_used_max_mb", "mem_used_max_pct",
    "net_rx_avg_mb_s", "net_rx_max_mb_s", "net_tx_avg_mb_s", "net_tx_max_mb_s",
    "disk_read_avg_mb_s", "disk_read_max_mb_s",
    "disk_write_avg_mb_s", "disk_write_max_mb_s",
    "gpu_count", "gpu_util_avg_pct", "gpu_util_p95_pct", "gpu_util_max_pct",
    "gpu_mem_used_avg_mb", "gpu_mem_used_p95_mb", "gpu_mem_used_max_mb",
    "gpu_mem_total_mb", "gpu_mem_used_max_pct",
    "gpu_power_avg_w", "gpu_power_p95_w", "gpu_power_max_w",
    "gpu_temp_max_c",
]

RESOURCE_RESULT_COLUMN_LABELS = {
    "resource_monitor_available": "资源监控可用",
    "resource_sample_count": "资源采样数",
    "cpu_util_avg_pct": "CPU平均使用率(%)",
    "cpu_util_p95_pct": "CPU P95使用率(%)",
    "cpu_util_max_pct": "CPU最大使用率(%)",
    "mem_used_avg_mb": "内存平均使用量(MB)",
    "mem_used_p95_mb": "内存P95使用量(MB)",
    "mem_used_max_mb": "内存最大使用量(MB)",
    "mem_used_max_pct": "内存最大使用率(%)",
    "net_rx_avg_mb_s": "网络平均接收(MB/s)",
    "net_rx_max_mb_s": "网络最大接收(MB/s)",
    "net_tx_avg_mb_s": "网络平均发送(MB/s)",
    "net_tx_max_mb_s": "网络最大发送(MB/s)",
    "disk_read_avg_mb_s": "磁盘平均读取(MB/s)",
    "disk_read_max_mb_s": "磁盘最大读取(MB/s)",
    "disk_write_avg_mb_s": "磁盘平均写入(MB/s)",
    "disk_write_max_mb_s": "磁盘最大写入(MB/s)",
    "gpu_count": "GPU数量",
    "gpu_util_avg_pct": "GPU平均使用率(%)",
    "gpu_util_p95_pct": "GPU P95使用率(%)",
    "gpu_util_max_pct": "GPU最大使用率(%)",
    "gpu_mem_used_avg_mb": "GPU显存平均使用量(MB)",
    "gpu_mem_used_p95_mb": "GPU显存P95使用量(MB)",
    "gpu_mem_used_max_mb": "GPU显存最大使用量(MB)",
    "gpu_mem_total_mb": "GPU总显存(MB)",
    "gpu_mem_used_max_pct": "GPU显存最大使用率(%)",
    "gpu_power_avg_w": "GPU平均功耗(W)",
    "gpu_power_p95_w": "GPU P95功耗(W)",
    "gpu_power_max_w": "GPU最大功耗(W)",
    "gpu_temp_max_c": "GPU最高温度(C)",
}

_SYSTEM_SAMPLE_KEYS = [
    "cpu_util_pct",
    "mem_total_mb", "mem_used_mb", "mem_available_mb", "mem_used_pct",
    "net_rx_mb_s", "net_tx_mb_s",
    "disk_read_mb_s", "disk_write_mb_s",
]


@dataclass(frozen=True)
class ResourceReaders:
    proc_stat: Callable[[], str]
    meminfo: Callable[[], str]
    net_dev: Callable[[], str]
    diskstats: Callable[[], str]
    nvidia_smi: Callable[[], str]


def default_readers():
    return ResourceReaders(
        proc_stat=lambda: Path("/proc/stat").read_text(encoding="utf-8"),
        meminfo=lambda: Path("/proc/meminfo").read_text(encoding="utf-8"),
        net_dev=lambda: Path("/proc/net/dev").read_text(encoding="utf-8"),
        diskstats=lambda: Path("/proc/diskstats").read_text(encoding="utf-8"),
        nvidia_smi=lambda: subprocess.run(
            NVIDIA_SMI_QUERY,
            check=True,
            capture_output=True,
            text=True,
            timeout=10,
        ).stdout,
    )


class ResourceMonitor:
    def __init__(
        self,
        *,
        output_dir,
        interval_sec=1.0,
        enabled=True,
        backend="nvidia-smi",
        readers=None,
    ):
        self.output_dir = Path(output_dir)
        self.interval_sec = float(interval_sec)
        self.enabled = enabled
        self.backend = backend
        self.readers = readers or default_readers()
        self.samples = []
        self.gpu_samples = []
        self.error_count = 0
        self._started_at = None
        self._last_sample_at = None
        self._previous_cpu = None
        self._previous_net = None
        self._previous_disk = None
        self._thread = None
        self._stop = threading.Event()
        self._lock = threading.Lock()

    def start(self):
        if not self.enabled:
            return
        if self._thread is not None and self._thread.is_alive():
            return
        self._stop.clear()
        self.sample_once()
        self._thread = threading.Thread(
            target=self._run,
            name="resource-monitor",
            daemon=True,
        )
        self._thread.start()

    def _run(self):
        while not self._stop.wait(self.interval_sec):
            self.sample_once()

    def sample_once(self, now=None):
        if not self.enabled:
            return

        timestamp = time.time() if now is None else now
        with self._lock:
            if self._started_at is None:
                self._started_at = timestamp
            elapsed_since_previous = (
                self.interval_sec
                if self._last_sample_at is None
                else timestamp - self._last_sample_at
            )
            sample = {
                "timestamp": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime(timestamp)),
                "elapsed_s": _round(timestamp - self._started_at),
            }
            self._sample_system(sample, elapsed_since_previous)
            self._sample_gpu(sample)
            self.samples.append(sample)
            self._last_sample_at = timestamp

    def _sample_system(self, sample, elapsed_s):
        try:
            current_cpu = parse_proc_stat(self.readers.proc_stat())
            sample["cpu_util_pct"] = (
                cpu_utilization_pct(self._previous_cpu, current_cpu)
                if self._previous_cpu is not None
                else None
            )
            self._previous_cpu = current_cpu
        except Exception:
            self.error_count += 1

        try:
            sample.update(parse_meminfo(self.readers.meminfo()))
        except Exception:
            self.error_count += 1

        try:
            current_net = parse_net_dev(self.readers.net_dev())
            sample.update(
                network_rates_mb_s(self._previous_net, current_net, elapsed_s=elapsed_s)
                if self._previous_net is not None
                else {"net_rx_mb_s": None, "net_tx_mb_s": None}
            )
            self._previous_net = current_net
        except Exception:
            self.error_count += 1

        try:
            current_disk = parse_diskstats(self.readers.diskstats())
            sample.update(
                disk_rates_mb_s(self._previous_disk, current_disk, elapsed_s=elapsed_s)
                if self._previous_disk is not None
                else {"disk_read_mb_s": None, "disk_write_mb_s": None}
            )
            self._previous_disk = current_disk
        except Exception:
            self.error_count += 1

    def _sample_gpu(self, sample):
        if self.backend != "nvidia-smi":
            apply_gpu_aggregate(sample, [])
            return

        try:
            gpus = parse_nvidia_smi_csv(self.readers.nvidia_smi())
        except Exception:
            self.error_count += 1
            gpus = []

        self.gpu_samples.extend(gpus)
        apply_gpu_aggregate(sample, gpus)

    def stop(self):
        if not self.enabled:
            return {"available": False, "sample_count": 0, "aggregate": {}}

        was_started = self._thread is not None
        self._stop.set()
        if self._thread is not None:
            self._thread.join(timeout=max(self.interval_sec, 1.0) + 1.0)

        if was_started and self._started_at is not None:
            self.sample_once()

        with self._lock:
            samples = list(self.samples)
            gpu_samples = list(self.gpu_samples)
            error_count = self.error_count

        summary = summarize_samples(
            samples,
            gpu_details=summarize_gpus(gpu_samples),
            error_count=error_count,
            interval_sec=self.interval_sec,
            backend=self.backend,
        )
        write_samples_csv(self.output_dir / "resource_samples.csv", samples)
        write_summary_json(self.output_dir / "resource_summary.json", summary)
        return summary


def parse_proc_stat(text):
    for line in text.splitlines():
        parts = line.split()
        if not parts or parts[0] != "cpu":
            continue

        values = [int(value) for value in parts[1:]]
        idle = values[3] if len(values) > 3 else 0
        iowait = values[4] if len(values) > 4 else 0
        return {
            "total": sum(values),
            "idle": idle + iowait,
            "values": values,
        }

    raise ValueError("missing aggregate cpu line")


def cpu_utilization_pct(previous, current):
    total_delta = current["total"] - previous["total"]
    idle_delta = current["idle"] - previous["idle"]
    if total_delta <= 0 or idle_delta < 0 or idle_delta > total_delta:
        return None

    busy_delta = total_delta - idle_delta
    return busy_delta / total_delta * 100.0


def parse_meminfo(text):
    values_kb = {}
    for line in text.splitlines():
        if ":" not in line:
            continue
        key, raw_value = line.split(":", 1)
        parts = raw_value.strip().split()
        if not parts:
            continue
        values_kb[key] = float(parts[0])

    total_mb = values_kb.get("MemTotal", 0.0) / 1024.0
    available_kb = values_kb.get("MemAvailable", values_kb.get("MemFree", 0.0))
    available_mb = available_kb / 1024.0
    used_mb = max(0.0, total_mb - available_mb)
    used_pct = used_mb / total_mb * 100.0 if total_mb else 0.0

    return {
        "mem_total_mb": total_mb,
        "mem_available_mb": available_mb,
        "mem_used_mb": used_mb,
        "mem_used_pct": used_pct,
    }


def parse_net_dev(text):
    rx_bytes = 0
    tx_bytes = 0
    interfaces = {}

    for line in text.splitlines():
        if ":" not in line:
            continue

        iface, raw_stats = line.split(":", 1)
        iface = iface.strip()
        if iface == "lo":
            continue

        fields = raw_stats.split()
        if len(fields) < 16:
            continue

        rx = int(fields[0])
        tx = int(fields[8])
        interfaces[iface] = {"rx_bytes": rx, "tx_bytes": tx}
        rx_bytes += rx
        tx_bytes += tx

    return {
        "rx_bytes": rx_bytes,
        "tx_bytes": tx_bytes,
        "interfaces": interfaces,
    }


def network_rates_mb_s(previous, current, *, elapsed_s):
    if elapsed_s <= 0:
        return {"net_rx_mb_s": None, "net_tx_mb_s": None}

    rx_delta = max(0, current["rx_bytes"] - previous["rx_bytes"])
    tx_delta = max(0, current["tx_bytes"] - previous["tx_bytes"])
    return {
        "net_rx_mb_s": rx_delta / _MB / elapsed_s,
        "net_tx_mb_s": tx_delta / _MB / elapsed_s,
    }


def parse_diskstats(text):
    all_devices = {}
    main_devices = {}
    main_read_sectors = 0
    main_write_sectors = 0
    all_read_sectors = 0
    all_write_sectors = 0
    saw_main_device = False

    for line in text.splitlines():
        parts = line.split()
        if len(parts) < 10:
            continue

        name = parts[2]
        if (
            name.startswith("loop")
            or name.startswith("ram")
            or name.startswith("dm-")
            or name.startswith("md")
        ):
            continue

        read_sectors = int(parts[5])
        write_sectors = int(parts[9])
        device = {
            "read_sectors": read_sectors,
            "write_sectors": write_sectors,
        }
        all_devices[name] = device
        all_read_sectors += read_sectors
        all_write_sectors += write_sectors

        if not _is_partition_device(name):
            saw_main_device = True
            main_devices[name] = device
            main_read_sectors += read_sectors
            main_write_sectors += write_sectors

    return {
        "read_sectors": main_read_sectors if saw_main_device else all_read_sectors,
        "write_sectors": main_write_sectors if saw_main_device else all_write_sectors,
        "devices": main_devices if saw_main_device else all_devices,
    }


def disk_rates_mb_s(previous, current, *, elapsed_s):
    if elapsed_s <= 0:
        return {"disk_read_mb_s": None, "disk_write_mb_s": None}

    read_delta = max(0, current["read_sectors"] - previous["read_sectors"])
    write_delta = max(0, current["write_sectors"] - previous["write_sectors"])
    return {
        "disk_read_mb_s": read_delta * _SECTOR_BYTES / _MB / elapsed_s,
        "disk_write_mb_s": write_delta * _SECTOR_BYTES / _MB / elapsed_s,
    }


def parse_nvidia_smi_csv(text):
    gpus = []
    for row in csv.reader(StringIO(text)):
        if not row or not any(cell.strip() for cell in row):
            continue

        padded = [cell.strip() for cell in row] + [""] * 8
        gpu_index, name, uuid, util, mem_used, mem_total, power, temperature = padded[:8]
        mem_used_mb = _parse_float(mem_used)
        mem_total_mb = _parse_float(mem_total)
        mem_used_pct = (
            mem_used_mb / mem_total_mb * 100.0
            if mem_used_mb is not None and mem_total_mb
            else None
        )

        gpus.append({
            "gpu_index": int(gpu_index),
            "gpu_name": name,
            "gpu_uuid": uuid,
            "gpu_util_pct": _parse_float(util),
            "gpu_mem_used_mb": mem_used_mb,
            "gpu_mem_total_mb": mem_total_mb,
            "gpu_mem_used_pct": mem_used_pct,
            "gpu_power_w": _parse_float(power),
            "gpu_temperature_c": _parse_float(temperature),
        })

    return gpus


def apply_gpu_aggregate(sample, gpus):
    gpus = list(gpus)
    sample["gpu_count"] = len(gpus)
    sample["gpu_util_avg_pct"] = _average(_gpu_values(gpus, "gpu_util_pct"))
    sample["gpu_util_max_pct"] = _maximum(_gpu_values(gpus, "gpu_util_pct"))
    sample["gpu_mem_used_avg_mb"] = _average(_gpu_values(gpus, "gpu_mem_used_mb"))
    sample["gpu_mem_used_max_mb"] = _maximum(_gpu_values(gpus, "gpu_mem_used_mb"))
    sample["gpu_mem_total_mb"] = _maximum(_gpu_values(gpus, "gpu_mem_total_mb"))
    sample["gpu_mem_used_max_pct"] = _maximum(_gpu_values(gpus, "gpu_mem_used_pct"))
    sample["gpu_power_avg_w"] = _average(_gpu_values(gpus, "gpu_power_w"))
    sample["gpu_power_max_w"] = _maximum(_gpu_values(gpus, "gpu_power_w"))
    sample["gpu_temp_max_c"] = _maximum(_gpu_values(gpus, "gpu_temperature_c"))
    sample["gpu_temperature_c"] = sample["gpu_temp_max_c"]


def summarize_gpus(gpu_samples):
    grouped = {}
    for gpu in gpu_samples:
        grouped.setdefault(gpu["gpu_index"], []).append(gpu)

    summaries = []
    for gpu_index in sorted(grouped):
        samples = grouped[gpu_index]
        first = samples[0]
        summaries.append({
            "gpu_index": gpu_index,
            "gpu_name": first.get("gpu_name"),
            "gpu_uuid": first.get("gpu_uuid"),
            "sample_count": len(samples),
            "gpu_util_avg_pct": _average(_gpu_values(samples, "gpu_util_pct")),
            "gpu_util_p95_pct": _p95(_gpu_values(samples, "gpu_util_pct")),
            "gpu_util_max_pct": _maximum(_gpu_values(samples, "gpu_util_pct")),
            "gpu_mem_used_avg_mb": _average(_gpu_values(samples, "gpu_mem_used_mb")),
            "gpu_mem_used_p95_mb": _p95(_gpu_values(samples, "gpu_mem_used_mb")),
            "gpu_mem_used_max_mb": _maximum(_gpu_values(samples, "gpu_mem_used_mb")),
            "gpu_mem_total_mb": _maximum(_gpu_values(samples, "gpu_mem_total_mb")),
            "gpu_mem_used_max_pct": _maximum(_gpu_values(samples, "gpu_mem_used_pct")),
            "gpu_power_avg_w": _average(_gpu_values(samples, "gpu_power_w")),
            "gpu_power_p95_w": _p95(_gpu_values(samples, "gpu_power_w")),
            "gpu_power_max_w": _maximum(_gpu_values(samples, "gpu_power_w")),
            "gpu_temp_avg_c": _average(_gpu_values(samples, "gpu_temperature_c")),
            "gpu_temp_p95_c": _p95(_gpu_values(samples, "gpu_temperature_c")),
            "gpu_temp_max_c": _maximum(_gpu_values(samples, "gpu_temperature_c")),
        })

    return summaries


def summarize_samples(
    samples,
    *,
    gpu_details,
    error_count=0,
    interval_sec=1.0,
    backend="nvidia-smi",
):
    sample_count = len(samples)
    gpu_details = list(gpu_details)
    aggregate = {}
    for sample_key, avg_key, p95_key, max_key in (
        ("cpu_util_pct", "cpu_util_avg_pct", "cpu_util_p95_pct", "cpu_util_max_pct"),
        ("mem_used_mb", "mem_used_avg_mb", "mem_used_p95_mb", "mem_used_max_mb"),
        ("mem_used_pct", "mem_used_avg_pct", "mem_used_p95_pct", "mem_used_max_pct"),
        ("net_rx_mb_s", "net_rx_avg_mb_s", "net_rx_p95_mb_s", "net_rx_max_mb_s"),
        ("net_tx_mb_s", "net_tx_avg_mb_s", "net_tx_p95_mb_s", "net_tx_max_mb_s"),
        ("disk_read_mb_s", "disk_read_avg_mb_s", "disk_read_p95_mb_s", "disk_read_max_mb_s"),
        (
            "disk_write_mb_s",
            "disk_write_avg_mb_s",
            "disk_write_p95_mb_s",
            "disk_write_max_mb_s",
        ),
    ):
        values = _sample_values(samples, sample_key)
        aggregate[avg_key] = _average(values)
        aggregate[p95_key] = _p95(values)
        aggregate[max_key] = _maximum(values)

    _aggregate_split_sample_values(
        aggregate,
        samples,
        avg_sample_key="gpu_util_avg_pct",
        max_sample_key="gpu_util_max_pct",
        fallback_key="gpu_util_avg_pct",
        avg_key="gpu_util_avg_pct",
        p95_key="gpu_util_p95_pct",
        max_key="gpu_util_max_pct",
    )
    _aggregate_split_sample_values(
        aggregate,
        samples,
        avg_sample_key="gpu_mem_used_avg_mb",
        max_sample_key="gpu_mem_used_max_mb",
        fallback_key="gpu_mem_used_mb",
        avg_key="gpu_mem_used_avg_mb",
        p95_key="gpu_mem_used_p95_mb",
        max_key="gpu_mem_used_max_mb",
    )
    _aggregate_split_sample_values(
        aggregate,
        samples,
        avg_sample_key="gpu_power_avg_w",
        max_sample_key="gpu_power_max_w",
        fallback_key="gpu_power_w",
        avg_key="gpu_power_avg_w",
        p95_key="gpu_power_p95_w",
        max_key="gpu_power_max_w",
    )
    _aggregate_split_sample_values(
        aggregate,
        samples,
        avg_sample_key="gpu_temp_max_c",
        max_sample_key="gpu_temp_max_c",
        fallback_key="gpu_temperature_c",
        avg_key="gpu_temp_avg_c",
        p95_key="gpu_temp_p95_c",
        max_key="gpu_temp_max_c",
    )

    aggregate["gpu_count"] = len(gpu_details)
    aggregate["gpu_mem_total_mb"] = _maximum(_sample_values(samples, "gpu_mem_total_mb"))
    aggregate["gpu_mem_used_max_pct"] = _maximum(
        _sample_values_or_fallback(samples, "gpu_mem_used_max_pct", "gpu_mem_used_pct")
    )

    system_available = _has_system_sample(samples)
    gpu_available = bool(gpu_details)
    return {
        "available": system_available or gpu_available,
        "system_available": system_available,
        "gpu_available": gpu_available,
        "backend": backend,
        "interval_sec": interval_sec,
        "sample_count": sample_count,
        "error_count": error_count,
        "aggregate": aggregate,
        "gpus": gpu_details,
    }


def write_samples_csv(path, samples):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=SAMPLE_HEADERS,
            extrasaction="ignore",
        )
        writer.writeheader()
        for sample in samples:
            writer.writerow({key: sample.get(key) for key in SAMPLE_HEADERS})


def write_summary_json(path, summary):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def flatten_summary_for_result(summary):
    aggregate = summary.get("aggregate", {})
    values = {}
    for column in RESOURCE_RESULT_COLUMNS:
        if column == "resource_monitor_available":
            values[column] = _result_value(summary.get("available", False))
        elif column == "resource_sample_count":
            values[column] = _result_value(summary.get("sample_count", 0))
        else:
            values[column] = _result_value(aggregate.get(column))
    return values


def prefixed_resource_columns(prefix):
    return [f"{prefix}_{column}" for column in RESOURCE_RESULT_COLUMNS]


def flatten_prefixed_summaries(summaries):
    values = {}
    for prefix, summary in summaries.items():
        flattened = flatten_summary_for_result(summary)
        for column in RESOURCE_RESULT_COLUMNS:
            values[f"{prefix}_{column}"] = flattened.get(column, "")
    return values


def append_summary_to_result_files(output_dir, summary):
    output_dir = Path(output_dir)
    values = flatten_summary_for_result(summary)

    csv_path = output_dir / "result.csv"
    if csv_path.exists():
        append_summary_to_csv(csv_path, values)

    xlsx_path = output_dir / "result.xlsx"
    if xlsx_path.exists():
        append_summary_to_xlsx(xlsx_path, values)


def append_prefixed_summaries_to_result_files(output_dir, summaries):
    output_dir = Path(output_dir)
    values = flatten_prefixed_summaries(summaries)
    if not values:
        return

    columns = list(values)
    csv_path = output_dir / "result.csv"
    if csv_path.exists():
        append_values_to_csv(csv_path, values, columns)

    xlsx_path = output_dir / "result.xlsx"
    if xlsx_path.exists():
        append_values_to_xlsx(xlsx_path, values, columns)


def append_summary_to_csv(path, values):
    append_values_to_csv(path, values, RESOURCE_RESULT_COLUMNS)


def append_values_to_csv(path, values, columns):
    path = Path(path)
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)

    merged_fieldnames = fieldnames + [
        column for column in columns
        if column not in fieldnames
    ]
    for row in rows:
        for column in columns:
            row[column] = _result_value(values.get(column, ""))

    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=merged_fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def append_summary_to_xlsx(path, values):
    append_values_to_xlsx(
        path,
        values,
        RESOURCE_RESULT_COLUMNS,
        column_labels=RESOURCE_RESULT_COLUMN_LABELS,
    )


def append_values_to_xlsx(path, values, columns, column_labels=None):
    try:
        import openpyxl
    except ImportError:
        return

    path = Path(path)
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    column_by_name = {
        worksheet.cell(row=1, column=column).value: column
        for column in range(1, worksheet.max_column + 1)
        if worksheet.cell(row=1, column=column).value is not None
    }

    next_column = worksheet.max_column + 1
    for column_name in columns:
        column = column_by_name.get(column_name)
        if column is None:
            column = next_column
            next_column += 1
            column_by_name[column_name] = column
            worksheet.cell(row=1, column=column, value=column_name)
        if worksheet.max_row >= 2:
            label = (
                column_labels.get(column_name, column_name)
                if column_labels is not None
                else column_name
            )
            worksheet.cell(
                row=2,
                column=column,
                value=label,
            )

    data_start_row = 3 if worksheet.max_row >= 2 else 2
    for row in range(data_start_row, worksheet.max_row + 1):
        for column_name in columns:
            worksheet.cell(
                row=row,
                column=column_by_name[column_name],
                value=_result_value(values.get(column_name, "")),
            )

    workbook.save(path)


def _is_partition_device(name):
    if name.startswith("dm-") or name.startswith("md"):
        return False
    if name.startswith("nvme") or name.startswith("mmcblk"):
        head, sep, tail = name.rpartition("p")
        return bool(head and sep and tail.isdigit())
    return name[-1:].isdigit()


def _gpu_values(gpus, key):
    return [
        float(gpu[key])
        for gpu in gpus
        if gpu.get(key) is not None
    ]


def _aggregate_split_sample_values(
    aggregate,
    samples,
    *,
    avg_sample_key,
    max_sample_key,
    fallback_key,
    avg_key,
    p95_key,
    max_key,
):
    avg_values = _sample_values_or_fallback(samples, avg_sample_key, fallback_key)
    max_values = _sample_values_or_fallback(samples, max_sample_key, fallback_key)
    aggregate[avg_key] = _average(avg_values)
    aggregate[p95_key] = _p95(avg_values)
    aggregate[max_key] = _maximum(max_values)


def _sample_values_or_fallback(samples, primary_key, fallback_key):
    values = _sample_values(samples, primary_key)
    if values:
        return values
    return _sample_values(samples, fallback_key)


def _has_system_sample(samples):
    return any(
        sample.get(key) is not None
        for sample in samples
        for key in _SYSTEM_SAMPLE_KEYS
    )


def _result_value(value):
    if isinstance(value, bool):
        return "true" if value else "false"
    if value is None:
        return ""
    return value


def _round(value):
    return round(value, 6)


def _parse_float(value):
    value = value.strip()
    if not value or value.upper() == "N/A" or value.lower() == "not supported" or value.startswith("["):
        return None
    return float(value)


def _sample_values(samples, key):
    return [
        float(sample[key])
        for sample in samples
        if sample.get(key) is not None
    ]


def _average(values):
    if not values:
        return None
    return sum(values) / len(values)


def _maximum(values):
    if not values:
        return None
    return max(values)


def _p95(values):
    if not values:
        return None
    sorted_values = sorted(values)
    index = math.ceil(0.95 * len(sorted_values)) - 1
    return sorted_values[index]
